"""
ITU-R BS.1770-5 响度测量仪 v3.2 (整合修复版)
修复：
- 测量算法（process_audio 变量定义、最大值追踪）
- 智能声道匹配（支持点号分隔的声道标识）
- 术语统一（节目响度、最大短时/瞬时响度）
- ADM解析兼容性（支持旧版adm_parser）
"""

import os
import sys
import time
import re
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, Optional, List, Tuple
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QProgressBar, QTableWidget,
    QTableWidgetItem, QGroupBox, QMessageBox, QComboBox, QDialog,
    QListWidget, QAbstractItemView, QDialogButtonBox, QLineEdit,
    QFrame, QCheckBox, QSpinBox, QTextEdit, QSizePolicy,
    QButtonGroup, QScrollArea, QGridLayout, QStyledItemDelegate, QHeaderView
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QFont, QColor


@dataclass
class LoudnessStandard:
    """响度标准配置"""
    name: str
    integrated_target: float
    integrated_tolerance: float
    true_peak_limit: float
    
    def check_integrated(self, value: float) -> bool:
        """检查节目响度是否符合标准"""
        return abs(value - self.integrated_target) <= self.integrated_tolerance
    
    def check_true_peak(self, value: float) -> bool:
        """检查真峰值是否符合标准"""
        return value <= self.true_peak_limit


# 响度标准定义
LOUDNESS_STANDARDS = {
    "GY/T 282-2014 (中国广电-电视)": LoudnessStandard(
        name="GY/T 282-2014", integrated_target=-24.0, integrated_tolerance=2.0, true_peak_limit=-2.0
    ),
    "GY/T 377-2023 (中国广电-网络/嘈杂环境)": LoudnessStandard(
        name="GY/T 377-2023", integrated_target=-15.0, integrated_tolerance=2.0, true_peak_limit=-2.0
    ),
    "EBU R128 (欧洲广播)": LoudnessStandard(
        name="EBU R128", integrated_target=-23.0, integrated_tolerance=1.0, true_peak_limit=-1.0
    ),
    "ATSC A/85 (美国电视)": LoudnessStandard(
        name="ATSC A/85", integrated_target=-24.0, integrated_tolerance=1.0, true_peak_limit=-1.0
    ),
}


from renderers.ear_renderer import render_adm, get_supported_layouts, get_adm_info
from mono_channel_matcher import (
    SmartMultiMonoDialog,
    CHANNEL_TEMPLATES,
    auto_match_mono_files,
)


class LoudnessCurveWidget(QWidget):
    """短时响度/时间曲线图（QPainter 自绘）"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(240)
        self.setMaximumHeight(280)
        self.values: list = []
        self.duration: float = 0.0
        self.target_lufs: float = -23.0
        self.y_min: float = -40.0
        self.y_max: float = 0.0

    def set_data(self, values, duration, target_lufs):
        self.values = list(values) if values else []
        self.duration = float(duration) if duration else 0.0
        self.target_lufs = float(target_lufs)
        self.update()

    def paintEvent(self, event):
        from PySide6.QtGui import QPainter, QPen, QColor
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        w, h = self.width(), self.height()
        margin_left, margin_right = 34, 15
        margin_top, margin_bottom = 15, 42
        chart_w = w - margin_left - margin_right
        chart_h = h - margin_top - margin_bottom

        # 背景（60% 透明度）
        painter.fillRect(self.rect(), QColor(26, 26, 46, 153))
        painter.fillRect(margin_left, margin_top, chart_w, chart_h, QColor(22, 33, 62, 153))

        if not self.values or self.duration <= 0:
            painter.setPen(QColor("#888"))
            painter.drawText(self.rect(), Qt.AlignCenter, self.tr("等待测量数据..."))
            painter.end()
            return

        # Y 轴网格线与刻度
        y_steps = 4  # -40, -30, -20, -10, 0
        for i in range(y_steps + 1):
            y_val = self.y_min + i * (self.y_max - self.y_min) / y_steps
            y_pos = margin_top + chart_h - (y_val - self.y_min) / (self.y_max - self.y_min) * chart_h
            painter.setPen(QPen(QColor("#333333"), 1, Qt.DashLine))
            painter.drawLine(margin_left, int(y_pos), margin_left + chart_w, int(y_pos))
            painter.setPen(QColor("#aaa"))
            painter.drawText(2, int(y_pos) - 6, 30, 12, Qt.AlignRight | Qt.AlignVCenter, f"{int(y_val)}")

        # 目标响度值在纵轴上额外标出（橙色）
        if self.y_min <= self.target_lufs <= self.y_max:
            y_target_pos = margin_top + chart_h - (self.target_lufs - self.y_min) / (self.y_max - self.y_min) * chart_h
            painter.setPen(QColor("#f39c12"))
            painter.drawText(2, int(y_target_pos) - 6, 30, 12, Qt.AlignRight | Qt.AlignVCenter,
                             f"{self.target_lufs:.0f}")

        # X 轴网格线与刻度
        x_ticks = self._calc_x_ticks(self.duration)
        for t in x_ticks:
            x_pos = margin_left + t / self.duration * chart_w
            painter.setPen(QPen(QColor("#333333"), 1, Qt.DashLine))
            painter.drawLine(int(x_pos), margin_top, int(x_pos), margin_top + chart_h)
            painter.setPen(QColor("#aaa"))
            painter.drawText(int(x_pos) - 20, margin_top + chart_h + 2, 40, 12, Qt.AlignCenter, f"{t}s")

        # 目标响度水平线
        if self.y_min <= self.target_lufs <= self.y_max:
            y_target = margin_top + chart_h - (self.target_lufs - self.y_min) / (self.y_max - self.y_min) * chart_h
            painter.setPen(QPen(QColor("#f39c12"), 1.5, Qt.DashLine))
            painter.drawLine(margin_left, int(y_target), margin_left + chart_w, int(y_target))
            painter.setPen(QColor("#f39c12"))
            painter.drawText(margin_left + chart_w - 55, int(y_target) - 14, 55, 12, Qt.AlignRight,
                             f"Target {self.target_lufs:.0f}")

        # 折线
        pen = QPen(QColor("#3498db"), 1.5)
        painter.setPen(pen)
        n = len(self.values)
        points = []
        for i, v in enumerate(self.values):
            if v == float('-inf') or v < self.y_min:
                v = self.y_min
            elif v > self.y_max:
                v = self.y_max
            x = margin_left + (i / max(self.duration, 1)) * chart_w
            y = margin_top + chart_h - (v - self.y_min) / (self.y_max - self.y_min) * chart_h
            points.append((x, y))

        for i in range(len(points) - 1):
            painter.drawLine(int(points[i][0]), int(points[i][1]),
                             int(points[i + 1][0]), int(points[i + 1][1]))

        # 轴标签
        painter.setPen(QColor("#ccc"))
        painter.drawText(0, h // 2 - 40, 6, 80, Qt.AlignCenter, "L\nK\nF\nS")
        painter.drawText(w // 2 - 30, h - 22, 60, 12, Qt.AlignCenter, self.tr("时间 (s)"))

        painter.end()

    def _calc_x_ticks(self, duration):
        if duration <= 60:
            step = 10
        elif duration <= 300:
            step = 30
        elif duration <= 600:
            step = 60
        else:
            step = 120
        ticks = [0]
        t = step
        while t < duration:
            ticks.append(t)
            t += step
        if duration > 0 and (not ticks or abs(ticks[-1] - duration) > 3):
            ticks.append(int(duration))
        return ticks


class ExportOptionsDialog(QDialog):
    def __init__(self, has_detailed, parent=None):
        super().__init__(parent)
        self.setWindowTitle(self.tr("导出选项"))
        self.setMinimumSize(400, 300)
        
        self.has_detailed = has_detailed
        
        layout = QVBoxLayout(self)
        
        layout.addWidget(QLabel(self.tr("导出格式:")))
        self.format_combo = QComboBox()
        self.format_combo.addItems([self.tr("TXT (文本报告)"), self.tr("JSON (结构化数据)"), self.tr("CSV (表格数据)")])
        layout.addWidget(self.format_combo)
        
        layout.addSpacing(20)
        layout.addWidget(QLabel(self.tr("详细程度:")))
        
        self.summary_check = QCheckBox(self.tr("总体概况 (节目响度/最大短时/最大瞬时/真峰值/LRA)"))
        self.summary_check.setChecked(True)
        layout.addWidget(self.summary_check)
        
        layout.addWidget(QLabel(self.tr("Excel导出将包含:\n• 整体测量结果\n• 每秒短时响度 (3秒滑动窗口)\n• 每秒最大真峰值\n• 超标数值以红色标注")))
        
        layout.addStretch()
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def on_detailed_changed(self, state):
        self.interval_spin.setEnabled(state == Qt.Checked)
    
    def get_options(self):
        return {
            'format': ['txt', 'json', 'csv'][self.format_combo.currentIndex()],
            'include_summary': self.summary_check.isChecked(),
            'include_detailed': True,
            'interval_ms': 1000
        }


class DetailedMeasurementWorker(QThread):
    progress_detail = Signal(int, float, float, float, str)
    finished_signal = Signal(dict)
    error = Signal(str)
    sub_step = Signal(str, int)
    
    def __init__(self, input_mode, input_data, config_name, standard):
        super().__init__()
        self.input_mode = input_mode
        self.input_data = input_data
        self.config_name = config_name
        self.standard = standard
        self.process_start_time = None  # 处理开始时间
        self.audio_duration = 0.0  # 音频总时长（秒）
    
    def run(self):
        try:
            start_time = time.time()
            sr = None
            num_channels = None
            total_samples = 0
            filename = self.tr("测量文件")
            
            # === 阶段1: 准备 (0-15%) ===
            self.sub_step.emit(self.tr("准备音频..."), 0)
            
            if self.input_mode == 'file':
                file_path = self.input_data
                info = sf.info(file_path)
                sr = info.samplerate
                num_channels = info.channels
                total_samples = info.frames
                filename = Path(file_path).name
                file_size = Path(file_path).stat().st_size
                
                # 所有文件统一走流式读取，避免小文件/大文件逻辑分叉
                audio_iter = self._iter_audio_file(file_path, file_size)
                    
            elif self.input_mode == 'adm':
                parser = self.input_data
                info = sf.info(parser.file_path)
                sr = info.samplerate
                num_channels = info.channels
                total_samples = info.frames
                filename = self.tr("ADM文件")
                
                # ADM 也使用流式读取
                audio_iter = parser.iter_audio_blocks(block_samples=sr, dtype='float32')
                
            elif self.input_mode == 'mono_list':
                mono_files = self.input_data
                sr, num_channels, total_samples, audio_iter = self._iter_mono_files(mono_files)
                filename = self.tr("多单声道文件")
            
            if sr is None or num_channels is None:
                raise ValueError(self.tr("音频加载失败"))
            
            actual_duration = total_samples / sr
            self.audio_duration = actual_duration  # 保存音频时长用于倍速计算
            
            # === 阶段2: 初始化测量器 (15-20%) ===
            self.sub_step.emit(self.tr("初始化: {num_channels} ch, {actual_duration:.1f} s").format(num_channels=num_channels, actual_duration=actual_duration), 15)
            
            if self.input_mode == 'adm':
                adm_config = self.input_data.adm.to_itu1770_config(num_channels)
                meter = ITU1770Meter(adm_config, sr)
            else:
                if self.config_name == "自动检测":
                    config = ITU1770Meter.auto_config(num_channels)
                else:
                    config_name = self.config_name if self.config_name in ITU1770Meter.CONFIGS else ITU1770Meter.auto_config(num_channels)
                    config = ITU1770Meter.CONFIGS.get(config_name, ITU1770Meter.auto_config(num_channels))
                
                # 防御性保护：配置声道数必须与音频通道数一致，否则回退到自动检测
                if len(config) != num_channels:
                    print(f"[测量器] 配置'{self.config_name}'({len(config)}ch)与音频{num_channels}ch不匹配，回退到自动检测")
                    config = ITU1770Meter.auto_config(num_channels)
                
                meter = ITU1770Meter(config, sr)
            
            meter.reset(sr)
            self.sub_step.emit(self.tr("开始测量..."), 20)
            self.process_start_time = time.time()  # 记录处理开始时间

            # === 阶段3: 响度测量 (20-90%) ===
            def on_process_progress(step_name, overall_progress_pct):
                # 检查是否请求中断
                if self.isInterruptionRequested():
                    raise RuntimeError("__INTERRUPTED__")

                # 20-90% 区间映射
                progress_pct = 20 + overall_progress_pct * 0.7

                # 计算实时处理倍速（按进度百分比推算已处理时长）
                speed_str = ""
                if self.process_start_time and self.audio_duration > 0:
                    elapsed = time.time() - self.process_start_time
                    processed_time = (overall_progress_pct / 100.0) * self.audio_duration
                    if elapsed > 0 and processed_time > 0:
                        speed_ratio = processed_time / elapsed
                        speed_str = self.tr(" | ⚡{ratio:.1f}x 实时").format(ratio=speed_ratio)

                self.sub_step.emit(f"{step_name}{speed_str}", int(progress_pct))

            try:
                processed_samples = 0
                for chunk, _sr in audio_iter:
                    if self.isInterruptionRequested():
                        raise RuntimeError("__INTERRUPTED__")
                    
                    if chunk.ndim == 1:
                        chunk = chunk.reshape(-1, 1)
                    
                    progress = (processed_samples / total_samples) * 100.0 if total_samples > 0 else 100.0
                    meter.feed(chunk, progress_callback=on_process_progress, overall_progress=progress)
                    processed_samples += chunk.shape[0]
                
                result = meter.finalize(progress_callback=on_process_progress)
            except RuntimeError as e:
                if str(e) == "__INTERRUPTED__":
                    self.sub_step.emit(self.tr("测量已停止"), 0)
                    return
                raise
            
            # === 阶段4: 最终计算 (90-95%) ===
            self.sub_step.emit(self.tr("计算最终指标..."), 90)
            
            # 获取结果（使用最大值）
            integrated = result['integrated']
            short_term = result['max_short_term'] if result['max_short_term'] != -np.inf else result['short_term']
            momentary = result['max_momentary'] if result['max_momentary'] != -np.inf else result['momentary']
            lra = result['lra']
            true_peak = result['true_peak']
            
            # 收集详细时序数据用于导出
            detailed_data = self._build_detailed_data(result, actual_duration)
            
            # === 阶段5: 完成 (95-100%) ===
            self.sub_step.emit(self.tr("整理结果..."), 95)
            
            final_results = {
                'integrated': integrated,
                'short_term': short_term,
                'momentary': momentary,
                'true_peak': true_peak,
                'lra': lra,
                'max_true_peak': true_peak,
                'duration': actual_duration,
                'filename': filename,
                'sample_rate': sr,
                'channels': num_channels,
                'processing_time': time.time() - start_time,
                'detailed_data': detailed_data,
                'short_term_curve': result.get('short_term_curve', []),
            }
            
            self.sub_step.emit(self.tr("完成"), 100)
            self.finished_signal.emit(final_results)
            
        except Exception as e:
            import traceback
            self.error.emit(f"{str(e)}\n{traceback.format_exc()}")
    
    def _iter_audio_file(self, file_path, file_size):
        """流式分块读取普通音频文件，显示进度 5-15%"""
        import soundfile as sf
        import numpy as np
        info = sf.info(file_path)
        sr = info.samplerate
        total_samples = info.frames
        
        with sf.SoundFile(file_path, 'r') as f:
            chunk_samples = sr  # 1 秒一块（按原始采样率）
            last_percent = -1
            for chunk in f.blocks(blocksize=chunk_samples, dtype='float32', always_2d=True):
                
                progress = 5 + (f.tell() / total_samples) * 10
                percent = int(progress)
                if percent != last_percent:
                    bytes_per_frame = file_size / total_samples if total_samples > 0 else 0
                    mb_loaded = (f.tell() * bytes_per_frame) / (1024 * 1024)
                    mb_total = file_size / (1024 * 1024)
                    self.sub_step.emit(self.tr("加载中... {mb_loaded:.1f}/{mb_total:.1f} MB").format(mb_loaded=mb_loaded, mb_total=mb_total), percent)
                    last_percent = percent
                yield np.ascontiguousarray(chunk), sr
    
    def _iter_mono_files(self, mono_files):
        """流式读取多个单声道文件，返回 (sr, num_channels, total_samples, generator)"""
        import soundfile as sf
        import numpy as np
        total_files = len(mono_files)
        
        # 分析所有文件
        max_samples = 0
        sr = None
        for path, name in mono_files:
            info = sf.info(path)
            max_samples = max(max_samples, info.frames)
            if sr is None:
                sr = info.samplerate
        
        num_channels = total_files
        self.sub_step.emit(self.tr("分析文件完成"), 7)
        
        def _generator():
            """内部生成器，负责实际分块读取"""
            import soundfile as sf
            import numpy as np
            handles = [sf.SoundFile(path, 'r') for path, _ in mono_files]
            try:
                chunk_samples = sr  # 1 秒一块
                chunk_idx = 0
                while True:
                    chunks = []
                    all_empty = True
                    for i, f in enumerate(handles):
                        chunk = f.read(chunk_samples, dtype='float32')
                        if len(chunk) > 0:
                            all_empty = False
                            if chunk.ndim > 1:
                                chunk = chunk[:, 0]
                        else:
                            chunk = np.zeros(0, dtype='float32')
                        chunks.append(chunk)
                    
                    if all_empty:
                        break
                    
                    max_len = max(len(c) for c in chunks)
                    if max_len == 0:
                        break
                    
                    out = np.ascontiguousarray(np.zeros((max_len, num_channels), dtype='float32'))
                    for i, chunk in enumerate(chunks):
                        out[:len(chunk), i] = chunk
                    
                    # 限制进度条刷新频率，避免与测量阶段的进度回调交替导致 UI 闪切
                    if chunk_idx % 10 == 0:
                        progress = 7 + (chunk_idx * chunk_samples / max_samples) * 8
                        self.sub_step.emit(self.tr("加载多单声道... {current}块").format(current=chunk_idx+1), int(progress))
                    chunk_idx += 1
                    
                    yield out, sr
            finally:
                for f in handles:
                    f.close()
        
        return sr, num_channels, max_samples, _generator()
    

    def _build_detailed_data(self, result: dict, duration: float) -> dict:
        """构建详细时序数据，精确到秒
        
        根据 ITU-R BS.1770-5 / EBU Tech 3341:
        - 短时响度: 3秒滑动窗口，每秒一个数据点
        - 真峰值: 4x过采样，每秒一个数据点
        """
        blocks = result.get('blocks', [])
        if not blocks:
            return None
        
        block_duration = 0.4  # 每个块400ms
        
        # === 1. 整体测量结果 ===
        summary = {
            'integrated_loudness': result.get('integrated', -np.inf),
            'max_short_term': result.get('max_short_term', -np.inf),
            'max_momentary': result.get('max_momentary', -np.inf),
            'max_true_peak': result.get('true_peak', -np.inf),
            'lra': result.get('lra', 0.0),
            'duration': duration
        }
        
        # === 2. 每秒短时响度 (3秒滑动窗口 = 7.5个块，取整为8个块) ===
        short_term_per_second = []
        window_blocks = 8  # 3秒窗口 ≈ 8个400ms块
        
        for second in range(int(duration) + 1):
            start_block = int(second / block_duration)
            end_block = min(start_block + window_blocks, len(blocks))
            
            if start_block < len(blocks):
                window = blocks[start_block:end_block]
                valid = [b for b in window if b > -70]
                if valid:
                    avg_loudness = float(np.mean(valid))
                    # 检查是否超标
                    is_exceed = False
                    if self.standard:
                        target = self.standard.integrated_target
                        tolerance = self.standard.integrated_tolerance
                        is_exceed = abs(avg_loudness - target) > tolerance
                    
                    short_term_per_second.append({
                        'time': float(second),
                        'lufs': round(avg_loudness, 2),
                        'is_exceed': is_exceed
                    })
        
        # === 3. 每秒最大真峰值 ===
        true_peak_per_second = []
        
        for second in range(int(duration) + 1):
            start_block = int(second / block_duration)
            end_block = min(start_block + int(1.0 / block_duration), len(blocks))
            
            if start_block < len(blocks):
                window = blocks[start_block:end_block]
                valid = [b for b in window if b > -70]
                if valid:
                    max_loudness = float(np.max(valid))
                    # 真峰值近似：使用瞬时响度 + 2dB 作为峰值估计
                    estimated_peak = max_loudness + 2.0
                    
                    # 检查是否超标
                    is_exceed = False
                    if self.standard:
                        tp_limit = self.standard.true_peak_limit
                        is_exceed = estimated_peak > tp_limit
                    
                    true_peak_per_second.append({
                        'time': float(second),
                        'dbtp': round(estimated_peak, 2),
                        'is_exceed': is_exceed
                    })
        
        return {
            'summary': summary,
            'short_term_per_second': short_term_per_second,
            'true_peak_per_second': true_peak_per_second
        }



class ADMRenderWorker(QThread):
    """ADM 渲染后台线程，带进度反馈"""
    progress = Signal(int)
    status = Signal(str)
    finished_signal = Signal(str)
    error = Signal(str)
    
    def __init__(self, input_path: str, target_layout: str, num_objects: int = 0):
        super().__init__()
        self.input_path = input_path
        self.target_layout = target_layout
        self.num_objects = num_objects
        self._cancelled = False
    
    def run(self):
        try:
            import sys
            from pathlib import Path
            src_dir = str(Path(__file__).parent)
            if src_dir not in sys.path:
                sys.path.insert(0, src_dir)
            print(f"[ADMRenderWorker] 开始导入 ear_renderer...")
            from renderers.ear_renderer import render_adm_with_progress
            print(f"[ADMRenderWorker] ear_renderer 导入完成")
            
            # EAR 在产出第一个数据块前需要大量初始化（解析 ADM、构建渲染图等），
            # 提前 emit 初始化状态，避免进度条长时间停在 0% 看起来像卡死。
            if self.num_objects > 0:
                init_msg = self.tr("正在初始化 ADM 渲染器（{n} 个对象）...").format(n=self.num_objects)
            else:
                init_msg = self.tr("正在初始化 ADM 渲染器...")
            self.status.emit(init_msg)
            
            def on_progress(percent: int):
                if not self._cancelled:
                    self.progress.emit(percent)
            
            print(f"[ADMRenderWorker] 调用 render_adm_with_progress: {self.input_path} -> {self.target_layout}")
            output_path = render_adm_with_progress(
                self.input_path,
                self.target_layout,
                progress_callback=on_progress,
            )
            print(f"[ADMRenderWorker] render_adm_with_progress 返回: {output_path}")
            self.finished_signal.emit(output_path)
        except Exception as e:
            import traceback
            err_detail = traceback.format_exc()
            print(f"[ADM渲染错误] {err_detail}")
            self.error.emit(str(e))
    
    def cancel(self):
        self._cancelled = True


class LoudnessMeterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("LouDuck")
        self.setMinimumSize(1280, 620)
        self.resize(1280, 650)
        
        self.current_results = None
        self.current_file = None
        self.mono_files = None
        self.current_standard = LOUDNESS_STANDARDS["GY/T 282-2014 (中国广电-电视)"]
        self.worker = None
        
        # 启用拖放
        self.setAcceptDrops(True)
        
        self._setup_ui()
        self._apply_theme()
    
    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        
        # 背景图片层（放在最底层，不拦截鼠标事件）
        self.bg_label = QLabel(central)
        from PySide6.QtGui import QPixmap
        bg_path = str(_get_resource_path("assets/bg.png"))
        self.bg_label.setPixmap(QPixmap(bg_path))
        self.bg_label.setScaledContents(True)
        self.bg_label.setAttribute(Qt.WA_TransparentForMouseEvents)
        self.bg_label.setGeometry(central.rect())
        self.bg_label.lower()
        
        main_layout = QHBoxLayout(central)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        left = self._create_left_panel()
        left.setMinimumWidth(420)
        main_layout.addWidget(left, 40)
        
        center = self._create_center_panel()
        center.setMaximumWidth(380)
        main_layout.addWidget(center, 25)
        
        right = self._create_right_panel()
        right.setMinimumWidth(260)
        main_layout.addWidget(right, 30)

    def _update_mono_files_list(self):
        """更新多单声道文件列表显示，支持编辑声道和顺序"""
        self.mono_files_table.blockSignals(True)
        if not self.mono_files:
            self.mono_files_table.setRowCount(0)
            self.mono_files_group.setTitle(self.tr('📋 声道匹配 (双击声道可编辑)'))
            self.mono_files_table.blockSignals(False)
            return

        self.mono_files_table.setRowCount(len(self.mono_files))
        for i, (path, ch_name) in enumerate(self.mono_files):
            # 序号
            num_item = QTableWidgetItem(str(i + 1))
            num_item.setTextAlignment(Qt.AlignCenter)
            num_item.setFlags(num_item.flags() & ~Qt.ItemIsEditable)
            self.mono_files_table.setItem(i, 0, num_item)

            # 声道（可编辑，有代理下拉框）
            ch_item = QTableWidgetItem(ch_name)
            ch_item.setTextAlignment(Qt.AlignCenter)
            if ch_name == '?':
                ch_item.setForeground(QColor('#e74c3c'))
            else:
                ch_item.setForeground(QColor('#27ae60'))
            self.mono_files_table.setItem(i, 1, ch_item)

            # 文件名
            filename = Path(path).name
            file_item = QTableWidgetItem(filename)
            file_item.setToolTip(str(path))
            file_item.setFlags(file_item.flags() & ~Qt.ItemIsEditable)
            self.mono_files_table.setItem(i, 2, file_item)

        self.mono_files_group.setVisible(True)
        matched = sum(1 for _, ch in self.mono_files if ch != '?')
        self.mono_files_group.setTitle(self.tr('📋 声道匹配 ({matched}/{total} 已匹配)').format(matched=matched, total=len(self.mono_files)))
        self.mono_files_table.blockSignals(False)
    def parse_and_display_adm(self, file_path: str):
        """解析ADM文件并在UI中显示详细信息"""
        try:
            self.adm_info.clear()
            self.adm_info.setPlainText(self.tr("正在解析 ADM..."))
            QApplication.processEvents()  # 立即更新UI
            
            parser = BW64Parser(file_path)
            adm = parser.parse()
            
            if not adm:
                self.adm_info.setPlainText(self.tr("[错误] 无法解析 ADM 元数据\n\n可能原因：\n1. 文件不是有效的 ADM/BW64 格式\n2. XML 命名空间不匹配"))
                return
            
            # 检查解析结果是否为空
            if not adm.channel_formats and not adm.programmes:
                self.adm_info.setPlainText(self.tr("[警告] ADM 元数据解析为空\n\n可能原因：\n1. 命名空间检测失败\n2. 文件不包含 ADM 数据"))
                return
            
            # 收集信息
            lines = []
            lines.append(self.tr("📦 文件: {name}").format(name=Path(file_path).name))
            lines.append("")
            
            # 节目信息
            if adm.programmes:
                prog_name = adm.programmes[0].get('name', 'N/A')
                lines.append(self.tr("🎬 节目: {name}").format(name=prog_name))
            
            # 内容统计
            content_count = len(adm.contents)
            object_count = len(adm.objects)
            lines.append(self.tr("📊 内容: {cc} 个 Content, {oc} 个 Object").format(cc=content_count, oc=object_count))
            
            # 声床分析
            direct_speakers = [ch for ch in adm.channel_formats if ch.type == 'DirectSpeakers']
            objects_ch = [ch for ch in adm.channel_formats if ch.type == 'Objects']
            
            lines.append("")
            lines.append(self.tr("🔊 声床配置 ({count} DirectSpeakers):").format(count=len(direct_speakers)))
            
            # 显示声床详情
            for i, ch in enumerate(direct_speakers):
                pos_str = ""
                if ch.position:
                    az = ch.position.get('azimuth', 0)
                    el = ch.position.get('elevation', 0)
                    pos_str = f"az={az:6.1f}° el={el:6.1f}°"
                elif ch.speaker_label:
                    az, el = adm._speaker_label_to_angles(ch.speaker_label)
                    pos_str = f"az={az:6.1f}° el={el:6.1f}° [{ch.speaker_label}]"
                
                lfe_mark = " [LFE]" if ch.is_lfe else ""
                lines.append(f"  Ch{i:2d}: {ch.name:20s} {pos_str}{lfe_mark}")
            
            # 检测Objects
            has_objects = bool(objects_ch)
            if has_objects:
                lines.append("")
                lines.append(self.tr("⚠️ 包含 {count} 个动态对象 (Object)").format(count=len(objects_ch)))
                self.atmos_render_group.setVisible(True)
                self.atmos_render_label.setText(
                    self.tr("⚠️ 检测到 {count} 个动态对象，可选择渲染到目标声道布局后，点击“渲染并测量”测量。\n"
                            "⚠️ 点击中间面板“开始测量”将仅测量声道响度，不包含对象。\n"
                            "⚠️ 本软件使用EAR（EBU ADM Renderer）作为渲染器，渲染结果可能与Dolby或Audio Vivid存在差异，渲染后的响度测量结果仅供参考。").format(count=len(objects_ch))
                )
            else:
                self.atmos_render_group.setVisible(False)
            
            # 配置识别 - 兼容新旧版adm_parser
            ch_count = len(direct_speakers)
            lines.append("")
            
            # 尝试使用新版 detect_configuration 方法
            detected = None
            confidence = 0.0
            description = ""
            
            if hasattr(adm, 'detect_configuration'):
                try:
                    detected, confidence, description = adm.detect_configuration()
                except Exception as e:
                    print(f"[DEBUG] detect_configuration 失败: {e}")
            
            if detected and detected != 'unknown':
                self.config_combo.setCurrentText(detected)
                lines.append(self.tr("🎯 自动识别为: {desc} ({ch_count} ch 声床, 置信度 {conf})").format(desc=self.tr(description), ch_count=ch_count, conf=f"{confidence:.0%}"))
            else:
                # 回退到数量映射
                cfg_map = {
                    2: 'stereo', 6: '5.1', 8: '7.1', 
                    10: '5.1.4', 12: '7.1.4', 16: '9.1.6'
                }
                fallback = cfg_map.get(ch_count, self.tr('未知'))
                if ch_count in cfg_map:
                    self.config_combo.setCurrentText(cfg_map[ch_count])
                else:
                    self.config_combo.setCurrentText(self.tr("自动检测"))
                lines.append(self.tr("⚠️ 基于数量识别: {fallback} ({ch_count} ch)").format(fallback=fallback, ch_count=ch_count))
                if detected == 'unknown':
                    lines.append(self.tr("   (特征识别失败，请手动确认)"))
            
            # 渲染器与创作软件信息（合并到 adm_info 中）
            lines.append("")
            lines.append("─" * 36)
            lines.append(self.tr("🎛️ 渲染器与创作软件信息"))
            lines.append("─" * 36)
            
            if adm.renderer_info:
                r_info = adm.renderer_info
                r_text = self.tr('🎚️ {name}').format(name=r_info.get('name', self.tr('未知渲染器')))
                if r_info.get('version'):
                    r_text += f" v{r_info['version']}"
                if r_info.get('coordinate_mode'):
                    r_text += f" [{r_info['coordinate_mode']}]"
                if r_info.get('uri'):
                    r_text += f"\n   URI: {r_info['uri']}"
                lines.append(r_text)
            else:
                lines.append(self.tr("🎚️ 未检测到渲染器信息"))
            
            if adm.authoring_info:
                a_info = adm.authoring_info
                if a_info.get('authoring_tool'):
                    a_text = f"🛠️ {a_info['authoring_tool']}"
                    if a_info.get('authoring_tool_version'):
                        a_text += f" v{a_info['authoring_tool_version']}"
                    lines.append(a_text)
                else:
                    lines.append(self.tr("🛠️ 未检测到创作软件"))
                
                if a_info.get('reference_layout'):
                    lines.append(self.tr('📐 参考布局: {layout}').format(layout=a_info['reference_layout']))
            else:
                lines.append(self.tr("🛠️ 未检测到创作软件信息"))
            
            # 显示到UI
            self.adm_info.setPlainText("\n".join(lines))
            
            # 保存parser供后续使用
            self.current_adm_parser = parser
            
        except Exception as e:
            error_msg = f"[解析错误] {str(e)}"
            self.adm_info.setPlainText(error_msg)
            import traceback
            print("=" * 60)
            print("ADM解析错误详情:")
            print(traceback.format_exc())
            print("=" * 60)

    
    def _create_left_panel(self):
        """左侧面板：输入方式、文件信息、模式专属区域（支持滚动）"""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")

        panel = QGroupBox(self.tr("📁 输入"))
        panel.setMinimumWidth(340)
        panel.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 1px solid #667eea;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 8px;
                padding-bottom: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #667eea;
            }
        """)
        layout = QVBoxLayout(panel)
        layout.setSpacing(12)
        layout.setContentsMargins(14, 14, 14, 14)

        # === 1. 文件导入按钮（EasyImport） + 帮助按钮 ===
        file_import_layout = QHBoxLayout()
        file_import_layout.setSpacing(8)
        file_import_layout.setContentsMargins(0, 0, 0, 0)

        self.btn_file = QPushButton(self.tr("📁 文件导入（EasyImport）"))
        self.btn_file.setCursor(Qt.PointingHandCursor)
        self.btn_file.clicked.connect(self.browse)
        self.btn_file.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_file.setStyleSheet("""
            QPushButton {
                background-color: #667eea;
                border: 2px solid #764ba2;
                padding: 8px 10px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                color: white;
            }
            QPushButton:hover { background-color: #764ba2; }
        """)
        file_import_layout.addWidget(self.btn_file, stretch=88)

        self.btn_easy_import_help = QPushButton("?")
        self.btn_easy_import_help.setCursor(Qt.PointingHandCursor)
        self.btn_easy_import_help.setToolTip(self.tr("EasyImport 说明"))
        self.btn_easy_import_help.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_easy_import_help.setMinimumWidth(28)
        self.btn_easy_import_help.setStyleSheet("""
            QPushButton {
                background-color: #4a5568;
                border: 1px solid #718096;
                border-radius: 4px;
                font-size: 13px;
                font-weight: bold;
                color: white;
                padding: 2px 4px;
            }
            QPushButton:hover { background-color: #718096; }
        """)
        self.btn_easy_import_help.clicked.connect(self.show_easy_import_help)
        file_import_layout.addWidget(self.btn_easy_import_help, stretch=10)

        layout.addLayout(file_import_layout)

        # === 2. 文件信息卡片 ===
        # === 2. 文件信息卡片（单个多声道/ADM 模式显示） ===
        self.file_info_group = QGroupBox(self.tr("文件信息"))
        self.file_info_group.setStyleSheet("""
            QGroupBox {
                font-size: 12px;
                border: 1px solid #3498db;
            }
            QGroupBox::title { color: #3498db; }
        """)
        file_layout = QVBoxLayout(self.file_info_group)
        file_layout.setSpacing(8)

        self.filename_label = QLabel(self.tr("未选择文件"))
        self.filename_label.setAlignment(Qt.AlignCenter)
        self.filename_label.setWordWrap(True)
        self.filename_label.setStyleSheet("""
            font-size: 15px;
            font-weight: bold;
            color: #667eea;
            padding: 4px;
        """)
        file_layout.addWidget(self.filename_label)

        self.path_label = QLabel("")
        self.path_label.setAlignment(Qt.AlignCenter)
        self.path_label.setWordWrap(True)
        self.path_label.setStyleSheet("""
            font-size: 10px;
            color: #888888;
            padding: 2px;
        """)
        file_layout.addWidget(self.path_label)

        # 元数据网格
        meta_widget = QWidget()
        meta_layout = QGridLayout(meta_widget)
        meta_layout.setSpacing(6)
        meta_layout.setContentsMargins(0, 4, 0, 4)
        meta_layout.setColumnStretch(1, 1)
        meta_layout.setColumnStretch(3, 1)

        self.file_meta_labels: Dict[str, QLabel] = {}
        meta_fields = [
            ('format', self.tr('格式')), ('channels', self.tr('声道')),
            ('samplerate', self.tr('采样率')), ('bit_depth', self.tr('位深')),
            ('duration', self.tr('时长')), ('file_size', self.tr('大小')),
        ]
        for i, (key, name) in enumerate(meta_fields):
            row, col = divmod(i, 2)
            lbl_name = QLabel(f"{name}:")
            lbl_name.setStyleSheet("color: #aaa; font-size: 11px;")
            lbl_val = QLabel("-")
            lbl_val.setStyleSheet("color: #eee; font-size: 11px; font-weight: bold;")
            self.file_meta_labels[key] = lbl_val
            meta_layout.addWidget(lbl_name, row, col * 2)
            meta_layout.addWidget(lbl_val, row, col * 2 + 1)
        file_layout.addWidget(meta_widget)

        # 声道顺序展示
        self.channel_order_label = QLabel(self.tr("声道顺序: -"))
        self.channel_order_label.setStyleSheet("color: #aaa; font-size: 11px; padding: 2px;")
        self.channel_order_label.setWordWrap(True)
        file_layout.addWidget(self.channel_order_label)

        layout.addWidget(self.file_info_group)

        # === 3. 模式专属区域 ===
        # -- 单个多声道 --
        self.standard_section = QWidget()
        standard_layout = QVBoxLayout(self.standard_section)
        standard_layout.setContentsMargins(0, 0, 0, 0)
        standard_layout.setSpacing(8)

        cfg_layout = QHBoxLayout()
        cfg_layout.addWidget(QLabel(self.tr("声道配置:")))
        self.config_combo = QComboBox()
        self.config_combo.addItems([self.tr("自动检测"), "stereo", "5.1", "7.1", "5.1.4", "7.1.2", "7.1.4"])
        self.config_combo.setStyleSheet("""
            QComboBox {
                background-color: rgba(15, 52, 96, 153);
                border: 1px solid #667eea;
                padding: 4px;
                font-size: 12px;
                color: #eee;
            }
        """)
        cfg_layout.addWidget(self.config_combo)
        standard_layout.addLayout(cfg_layout)
        standard_layout.addStretch()
        layout.addWidget(self.standard_section)

        # -- ADM/BW64 --
        self.adm_section = QWidget()
        adm_layout = QVBoxLayout(self.adm_section)
        adm_layout.setContentsMargins(0, 0, 0, 0)
        adm_layout.setSpacing(8)

        self.adm_info_group = QGroupBox(self.tr("ADM 信息"))
        self.adm_info_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #e74c3c;
                border-radius: 6px;
                margin-top: 6px;
                padding-top: 6px;
                font-weight: bold;
                font-size: 11px;
            }
            QGroupBox::title { color: #e74c3c; }
        """)
        adm_info_layout = QVBoxLayout(self.adm_info_group)
        adm_info_layout.setContentsMargins(6, 2, 6, 6)
        adm_info_layout.setSpacing(4)

        self.adm_info = QTextEdit()
        self.adm_info.setReadOnly(True)
        self.adm_info.setPlaceholderText(self.tr("ADM文件信息将显示在这里..."))
        self.adm_info.setMaximumHeight(160)
        self.adm_info.setStyleSheet("""
            QTextEdit {
                background-color: #16213e;
                border: none;
                color: #eee;
                font-family: Consolas, Monaco, monospace;
                font-size: 11px;
                padding: 4px;
            }
        """)
        adm_info_layout.addWidget(self.adm_info)
        adm_layout.addWidget(self.adm_info_group)

        # Dolby Atmos / Audio Vivid 渲染选项
        self.atmos_render_group = QGroupBox(self.tr("🎧 沉浸式音频渲染"))
        self.atmos_render_group.setVisible(False)
        self.atmos_render_group.setMinimumHeight(160)
        self.atmos_render_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #9b59b6;
                border-radius: 6px;
                margin-top: 6px;
                padding-top: 6px;
                font-weight: bold;
                font-size: 11px;
            }
            QGroupBox::title { color: #9b59b6; }
        """)
        atmos_render_layout = QVBoxLayout(self.atmos_render_group)
        atmos_render_layout.setSpacing(6)
        atmos_render_layout.setContentsMargins(10, 10, 10, 10)

        self.atmos_render_label = QLabel(self.tr("检测到动态对象音频，可选择渲染到目标声道布局后测量响度。\n注意：点击中间面板“开始测量”将仅测量声道响度，不包含对象。"))
        self.atmos_render_label.setStyleSheet("color: #bbb; font-size: 10px;")
        self.atmos_render_label.setWordWrap(True)
        atmos_render_layout.addWidget(self.atmos_render_label)

        atmos_ctrl_layout = QHBoxLayout()
        atmos_ctrl_layout.addWidget(QLabel(self.tr("目标布局:")))
        self.atmos_layout_combo = QComboBox()
        self.atmos_layout_combo.addItems(get_supported_layouts())
        self.atmos_layout_combo.setCurrentText("5.1.4 (10ch)")
        self.atmos_layout_combo.setStyleSheet("""
            QComboBox {
                background-color: rgba(15, 52, 96, 153);
                border: 1px solid #9b59b6;
                padding: 4px;
                font-size: 12px;
                color: #eee;
            }
        """)
        atmos_ctrl_layout.addWidget(self.atmos_layout_combo, 1)

        self.atmos_render_btn = QPushButton(self.tr("🎯 渲染并测量"))
        self.atmos_render_btn.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                border: none;
                padding: 5px 10px;
                border-radius: 4px;
                font-size: 11px;
                color: white;
            }
            QPushButton:hover { background-color: #8e44ad; }
        """)
        self.atmos_render_btn.clicked.connect(self._render_and_measure_adm)
        atmos_ctrl_layout.addWidget(self.atmos_render_btn)
        atmos_render_layout.addLayout(atmos_ctrl_layout)

        adm_layout.addWidget(self.atmos_render_group)
        adm_layout.addStretch()
        layout.addWidget(self.adm_section)

        # -- 多单声道 --
        self.mono_section = QWidget()
        mono_layout = QVBoxLayout(self.mono_section)
        mono_layout.setContentsMargins(0, 0, 0, 0)
        mono_layout.setSpacing(8)

        # 模板选择 + 控制按钮
        ctrl_layout = QHBoxLayout()
        ctrl_layout.setSpacing(6)
        ctrl_layout.addWidget(QLabel(self.tr("声道模板:")))
        self.mono_template_combo = QComboBox()
        self.mono_template_combo.addItems([
            self.tr("自动检测"), "Stereo (2.0)", "5.1 (6ch)", "7.1 (8ch)",
            "7.1.2 (10ch)", "5.1.4 (10ch)", "7.1.4 (12ch)", self.tr("自定义")
        ])
        self.mono_template_combo.setStyleSheet("""
            QComboBox {
                background-color: #0f3460;
                border: 1px solid #27ae60;
                padding: 4px;
                font-size: 12px;
                color: #eee;
            }
        """)
        ctrl_layout.addWidget(self.mono_template_combo, 1)

        self.mono_match_btn = QPushButton(self.tr("🎯 自动匹配"))
        self.mono_match_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                border: none;
                padding: 5px 10px;
                border-radius: 4px;
                font-size: 11px;
                color: white;
            }
            QPushButton:hover { background-color: #2ecc71; }
            QPushButton:disabled { background-color: #1e8449; color: #aaa; }
        """)
        self.mono_match_btn.clicked.connect(self._on_mono_auto_match)
        ctrl_layout.addWidget(self.mono_match_btn)

        self.mono_up_btn = QPushButton("⬆️")
        self.mono_up_btn.setToolTip(self.tr("上移"))
        self.mono_up_btn.setStyleSheet("QPushButton { padding: 5px 8px; font-size: 11px; }")
        self.mono_up_btn.clicked.connect(self._on_mono_move_up)
        ctrl_layout.addWidget(self.mono_up_btn)

        self.mono_down_btn = QPushButton("⬇️")
        self.mono_down_btn.setToolTip(self.tr("下移"))
        self.mono_down_btn.setStyleSheet("QPushButton { padding: 5px 8px; font-size: 11px; }")
        self.mono_down_btn.clicked.connect(self._on_mono_move_down)
        ctrl_layout.addWidget(self.mono_down_btn)

        self.mono_del_btn = QPushButton("🗑️")
        self.mono_del_btn.setToolTip(self.tr("删除选中"))
        self.mono_del_btn.setStyleSheet("QPushButton { padding: 5px 8px; font-size: 11px; }")
        self.mono_del_btn.clicked.connect(self._on_mono_delete)
        ctrl_layout.addWidget(self.mono_del_btn)

        self.mono_clear_btn = QPushButton(self.tr("清空"))
        self.mono_clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #c0392b;
                border: none;
                padding: 5px 10px;
                border-radius: 4px;
                font-size: 11px;
                color: white;
            }
            QPushButton:hover { background-color: #e74c3c; }
        """)
        self.mono_clear_btn.clicked.connect(self._clear_mono_files)
        ctrl_layout.addWidget(self.mono_clear_btn)
        mono_layout.addLayout(ctrl_layout)

        self.mono_files_group = QGroupBox(self.tr("📋 声道匹配 (双击声道可编辑)"))
        self.mono_files_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        self.mono_files_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #27ae60;
                border-radius: 6px;
                margin-top: 6px;
                padding-top: 6px;
                font-weight: bold;
                font-size: 11px;
            }
            QGroupBox::title { color: #27ae60; }
        """)
        mono_files_layout = QVBoxLayout(self.mono_files_group)
        mono_files_layout.setSpacing(4)
        mono_files_layout.setContentsMargins(10, 10, 10, 10)

        self.mono_files_table = QTableWidget(0, 3)
        self.mono_files_table.setHorizontalHeaderLabels([self.tr('#'), self.tr('声道'), self.tr('文件名')])
        self.mono_files_table.verticalHeader().setVisible(False)
        self.mono_files_table.horizontalHeader().setStretchLastSection(True)
        self.mono_files_table.setMinimumHeight(180)
        self.mono_files_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.mono_files_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.mono_files_table.setEditTriggers(QAbstractItemView.DoubleClicked)
        self.mono_files_table.setStyleSheet("""
            QTableWidget {
                background-color: #16213e;
                border: 1px solid #27ae60;
                font-size: 10px;
            }
            QHeaderView::section {
                background-color: #27ae60;
                color: white;
                padding: 4px;
                font-size: 10px;
            }
            QTableWidget::item {
                padding: 2px;
                color: #eee;
            }
            QTableWidget::item:selected {
                background-color: #1a4a7a;
            }
        """)
        self.mono_files_table.setColumnWidth(0, 30)
        self.mono_files_table.setColumnWidth(1, 60)
        # 声道列使用 QComboBox 代理
        class ChannelComboDelegate(QStyledItemDelegate):
            def __init__(self, parent=None):
                super().__init__(parent)
                self.channels = ['L', 'R', 'C', 'LFE', 'Ls', 'Rs', 'Lss', 'Rss',
                                 'Lrs', 'Rrs', 'Ltf', 'Rtf', 'Ltr', 'Rtr', 'Ltb', 'Rtb', '?']
            def createEditor(self, parent, option, index):
                editor = QComboBox(parent)
                editor.addItems(self.channels)
                return editor
            def setEditorData(self, editor, index):
                text = index.model().data(index, Qt.EditRole)
                idx = editor.findText(text)
                if idx >= 0:
                    editor.setCurrentIndex(idx)
            def setModelData(self, editor, model, index):
                model.setData(index, editor.currentText(), Qt.EditRole)
        self.mono_files_table.setItemDelegateForColumn(1, ChannelComboDelegate(self))
        self.mono_files_table.itemChanged.connect(self._on_mono_item_changed)
        mono_files_layout.addWidget(self.mono_files_table, 1)

        mono_layout.addWidget(self.mono_files_group, 1)
        mono_layout.addStretch()
        layout.addWidget(self.mono_section, 1)

        layout.addStretch(0)
        scroll.setWidget(panel)

        # 初始化显隐
        self._update_mode_ui('mono')

        return scroll

    def _update_mode_buttons(self, active_mode: str):
        """更新输入模式按钮样式（单按钮模式，无操作）"""
        pass

    def _update_mode_ui(self, mode: str):
        """根据模式切换左侧各区块显隐"""
        is_file_mode = (mode == 'file')
        is_mono = is_file_mode and bool(self.mono_files)
        self.standard_section.setVisible(is_file_mode and not is_mono)
        self.mono_section.setVisible(is_mono)
        # ADM 信息区域：只要有已解析的 ADM 数据就显示
        has_adm = hasattr(self, 'current_adm_parser') and self.current_adm_parser is not None
        self.adm_section.setVisible(has_adm)
        if hasattr(self, 'file_info_group'):
            self.file_info_group.setVisible(is_file_mode)

    # ---------- 多单声道控制方法 ----------

    def _on_mono_auto_match(self):
        """对当前已加载的文件重新执行自动匹配"""
        if not self.mono_files:
            QMessageBox.information(self, self.tr("提示"), self.tr("请先选择文件"))
            return
        template_map = {
            "Stereo (2.0)": "Stereo (2.0)",
            "5.1 (6ch)": "5.1 (6ch)",
            "7.1 (8ch)": "7.1 (8ch)",
            "7.1.2 (10ch)": "7.1.2 (10ch)",
            "5.1.4 (10ch)": "5.1.4 (10ch)",
            "7.1.4 (12ch)": "7.1.4 (12ch)",
        }
        combo_text = self.mono_template_combo.currentText()
        template_name = template_map.get(combo_text, None)
        file_paths = [p for p, _ in self.mono_files]
        self.mono_files = auto_match_mono_files(file_paths, template_name=template_name)
        self._update_mono_files_list()

    def _on_mono_move_up(self):
        row = self.mono_files_table.currentRow()
        if row <= 0:
            return
        self.mono_files[row], self.mono_files[row - 1] = self.mono_files[row - 1], self.mono_files[row]
        self._update_mono_files_list()
        self.mono_files_table.selectRow(row - 1)

    def _on_mono_move_down(self):
        row = self.mono_files_table.currentRow()
        if row < 0 or row >= len(self.mono_files) - 1:
            return
        self.mono_files[row], self.mono_files[row + 1] = self.mono_files[row + 1], self.mono_files[row]
        self._update_mono_files_list()
        self.mono_files_table.selectRow(row + 1)

    def _on_mono_delete(self):
        row = self.mono_files_table.currentRow()
        if row < 0 or row >= len(self.mono_files):
            return
        del self.mono_files[row]
        self._update_mono_files_list()
        if self.mono_files and row < len(self.mono_files):
            self.mono_files_table.selectRow(row)
        elif self.mono_files:
            self.mono_files_table.selectRow(len(self.mono_files) - 1)

    def _on_mono_item_changed(self, item):
        """表格编辑完成后同步到 self.mono_files"""
        if item.column() != 1:
            return
        row = item.row()
        if row < 0 or row >= len(self.mono_files):
            return
        new_ch = item.text()
        path, _ = self.mono_files[row]
        self.mono_files[row] = (path, new_ch)
        # 更新颜色
        if new_ch == '?':
            item.setForeground(QColor('#e74c3c'))
        else:
            item.setForeground(QColor('#27ae60'))
        matched = sum(1 for _, ch in self.mono_files if ch != '?')
        self.mono_files_group.setTitle(self.tr('📋 声道匹配 ({matched}/{total} 已匹配)').format(matched=matched, total=len(self.mono_files)))


    def _clear_mono_files(self):
        """清空多单声道文件列表"""
        self.mono_files = []
        self._update_mono_files_list()
        self.filename_label.setText(self.tr("未选择文件"))
        self.filename_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #667eea;")
        self.path_label.setText("")

    def _update_channel_order(self, path: str):
        """读取并显示文件内部声道顺序"""
        try:
            info = sf.info(path)
            channels = info.channels
            channel_names = []

            # 尝试从 soundfile extra_info 解析 Channel Mask
            try:
                with sf.SoundFile(path) as f:
                    extra = f.extra_info
                    import re
                    m = re.search(r'Channel Mask\s*:\s*0x[0-9A-Fa-f]+\s*\(([^)]+)\)', extra)
                    if m:
                        channel_names = [c.strip() for c in m.group(1).split(',')]
            except Exception:
                pass

            # 若解析失败，使用默认配置
            if not channel_names and channels > 0:
                default_map = {
                    1: ['M'],
                    2: ['L', 'R'],
                    6: ['L', 'R', 'C', 'LFE', 'Ls', 'Rs'],
                    8: ['L', 'R', 'C', 'LFE', 'Lss', 'Rss', 'Lrs', 'Rrs'],
                    10: ['L', 'R', 'C', 'LFE', 'Ls', 'Rs', 'Ltf', 'Rtf', 'Ltr', 'Rtr'],
                    12: ['L', 'R', 'C', 'LFE', 'Lss', 'Rss', 'Lrs', 'Rrs', 'Ltf', 'Rtf', 'Ltb', 'Rtb'],
                }
                channel_names = default_map.get(channels, [f'Ch{i+1}' for i in range(channels)])

            order_text = ' → '.join(channel_names[:channels])
            self.channel_order_label.setText(self.tr("声道顺序: {order}").format(order=order_text))
            self.channel_order_label.setStyleSheet("color: #27ae60; font-size: 11px; font-weight: bold; padding: 2px;")
        except Exception as e:
            print(f"[声道顺序读取失败] {e}")
            self.channel_order_label.setText(self.tr("声道顺序: -"))
            self.channel_order_label.setStyleSheet("color: #aaa; font-size: 11px; padding: 2px;")

    def _clear_file_metadata(self):
        """清空文件元数据显示"""
        for lbl in self.file_meta_labels.values():
            lbl.setText("-")
            lbl.setStyleSheet("color: #eee; font-size: 11px; font-weight: bold;")
        if hasattr(self, 'channel_order_label'):
            self.channel_order_label.setText(self.tr("声道顺序: -"))
            self.channel_order_label.setStyleSheet("color: #aaa; font-size: 11px; padding: 2px;")

    def _update_file_metadata(self, path: str):
        """读取并显示文件元数据"""
        try:
            info = sf.info(path)
            p = Path(path)
            size = p.stat().st_size
            size_str = self._format_file_size(size)

            self.file_meta_labels['format'].setText(p.suffix.lstrip('.').upper())
            self.file_meta_labels['channels'].setText(str(info.channels))
            self.file_meta_labels['samplerate'].setText(f"{info.samplerate} Hz")
            self.file_meta_labels['bit_depth'].setText(str(info.subtype_info))
            self.file_meta_labels['duration'].setText(self._format_duration(info.duration))
            self.file_meta_labels['file_size'].setText(size_str)
        except Exception as e:
            print(f"[元数据读取失败] {e}")
            self._clear_file_metadata()

    def _format_duration(self, seconds: float) -> str:
        """格式化时长：不足1分钟显示秒，超过1分钟显示 xx分xx秒"""
        if seconds >= 60.0:
            minutes = int(seconds // 60)
            secs = int(seconds % 60)
            millis = int((seconds % 1) * 1000)
            if millis > 0:
                return self.tr("{minutes}分{secs:02d}.{millis:03d}秒").format(minutes=minutes, secs=secs, millis=millis)
            return self.tr("{minutes}分{secs:02d}秒").format(minutes=minutes, secs=secs)
        return self.tr("{seconds:.2f}秒").format(seconds=seconds)

    @staticmethod
    def _format_file_size(size_bytes: int) -> str:
        """格式化文件大小"""
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.1f} KB"
        elif size_bytes < 1024 * 1024 * 1024:
            return f"{size_bytes / (1024 * 1024):.1f} MB"
        else:
            return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"

    def dragEnterEvent(self, event):
        """拖拽进入"""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        """拖拽放下"""
        urls = event.mimeData().urls()
        if not urls:
            return
        path = urls[0].toLocalFile()
        if not path:
            return
        self._load_file_by_path(path)

    def _load_file_by_path(self, path: str):
        """根据路径自动判断模式并加载文件（统一文件导入模式）"""
        p = Path(path)
        ext = p.suffix.lower()

        if ext not in ('.wav', '.flac', '.mp3', '.ogg', '.bw64', '.adm'):
            QMessageBox.warning(self, self.tr("不支持的文件"), self.tr("无法识别该文件类型:\n{name}").format(name=p.name))
            return

        try:
            info = sf.info(path)
        except Exception as e:
            QMessageBox.warning(self, self.tr("文件错误"), self.tr("无法读取文件:\n{err}").format(err=e))
            return

        self.on_input_mode_changed('file')
        self.current_file = path
        self.mono_files = None
        self.filename_label.setText(f"✓ {p.name}")
        self.filename_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #27ae60;")
        self.path_label.setText(str(p.parent))
        self._update_file_metadata(path)
        self._update_channel_order(path)

        # 自动检测 ADM 元数据
        if ext in ('.wav', '.bw64', '.adm'):
            try:
                if is_adm_file(path):
                    self.current_adm_parser = BW64Parser(path)
                    self.current_adm_parser.parse()
                    self.parse_and_display_adm(path)
            except Exception:
                pass

        cfg_map = {2: 'stereo', 6: '5.1', 8: '7.1', 10: '5.1.4', 12: '7.1.4'}
        if info.channels in cfg_map:
            self.config_combo.setCurrentText(cfg_map[info.channels])

    
    def _create_center_panel(self):
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setSpacing(8)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # === 标题栏 ===
        header_widget = QWidget()
        header_widget.setMaximumHeight(300)
        header_widget.setMinimumHeight(200)
        header_widget.setStyleSheet('QWidget { background-color: transparent; border: none; color: #eee; }')
        header_layout = QVBoxLayout(header_widget)
        header_layout.setContentsMargins(15, 8, 15, 8)
        header_layout.setSpacing(3)
        header_layout.setAlignment(Qt.AlignCenter)
        
        # Logo 图片
        from PySide6.QtGui import QPixmap
        logo_label = QLabel()
        logo_pixmap = QPixmap(str(_get_resource_path("assets/centerlogo.png")))
        if not logo_pixmap.isNull():
            logo_pixmap = logo_pixmap.scaled(280, 120, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            logo_label.setPixmap(logo_pixmap)
        logo_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(logo_label)
        
        # Immersive Loudness
        name_label = QLabel('Immersive audio file Loudness measure tool')
        name_label.setStyleSheet('color: #a0b4e8; font-size: 12px; font-weight: bold; font-family: "Segoe UI", "Microsoft YaHei", sans-serif; letter-spacing: 1px; border: none;')
        name_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(name_label)
        
        # 底部行
        bottom_layout = QHBoxLayout()
        bottom_layout.setSpacing(0)
        bottom_layout.setContentsMargins(0, 2, 0, 0)
        
        # 左侧三行功能描述
        left_info = QVBoxLayout()
        left_info.setSpacing(0)
        left_info.setAlignment(Qt.AlignLeft | Qt.AlignBottom)
        for line_text in ["File type auto-detection", "Channel auto-match", "ADM analysis and render", "Excel export"]:
            lbl = QLabel(line_text)
            lbl.setStyleSheet('color: #8899cc; font-size: 8px; font-family: "Segoe UI", sans-serif; border: none;')
            left_info.addWidget(lbl)
        bottom_layout.addLayout(left_info)
        
        bottom_layout.addStretch(1)
        
        # 右侧版本和版权
        right_info = QVBoxLayout()
        right_info.setSpacing(0)
        right_info.setAlignment(Qt.AlignRight | Qt.AlignBottom)
        
        version_label = QLabel('v1.0.3  (build 260630)')
        version_label.setStyleSheet('color: #667eea; font-size: 8px; border: none;')
        version_label.setAlignment(Qt.AlignRight)
        right_info.addWidget(version_label)
        
        copyright_label = QLabel('© 2026 YOYH All Rights Reserved')
        copyright_label.setStyleSheet('color: #888; font-size: 7px; border: none;')
        copyright_label.setAlignment(Qt.AlignRight)
        right_info.addWidget(copyright_label)
        
        bottom_layout.addLayout(right_info)
        header_layout.addLayout(bottom_layout)
        layout.addWidget(header_widget)
        
        # === 原有内容 ===
        content_widget = QGroupBox(self.tr("⚙️ 标准与进度"))
        content_layout = QVBoxLayout(content_widget)
        content_layout.setSpacing(8)
        
        content_layout.addWidget(QLabel(self.tr("响度标准:")))
        self.std_combo = QComboBox()
        self.std_combo.setMaximumWidth(360)
        self.std_combo.addItem(self.tr("GY/T 282-2014 (中国广电-电视)"), "GY/T 282-2014 (中国广电-电视)")
        self.std_combo.addItem(self.tr("GY/T 377-2023 (中国广电-网络/嘈杂环境)"), "GY/T 377-2023 (中国广电-网络/嘈杂环境)")
        self.std_combo.addItem(self.tr("EBU R128 (欧洲广播)"), "EBU R128 (欧洲广播)")
        self.std_combo.addItem(self.tr("ATSC A/85 (美国电视)"), "ATSC A/85 (美国电视)")
        self.std_combo.currentTextChanged.connect(self.on_std_changed)
        self.std_combo.setCurrentIndex(0)
        content_layout.addWidget(self.std_combo)
        
        self.std_info = QLabel()
        self.std_info.setStyleSheet('background-color: #16213e; padding: 8px; border-radius: 4px;')
        self.update_std_info()
        content_layout.addWidget(self.std_info)
        
        content_layout.addSpacing(10)
        
        # 渲染进度（ADM 渲染专用，与测量进度独立）
        self.render_step_label = QLabel('')
        self.render_step_label.setStyleSheet('color: #9b59b6; font-weight: bold;')
        self.render_step_label.setVisible(False)
        content_layout.addWidget(self.render_step_label)
        
        self.render_progress = QProgressBar()
        self.render_progress.setVisible(False)
        self.render_progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #9b59b6;
                border-radius: 4px;
                text-align: center;
                color: white;
            }
            QProgressBar::chunk {
                background-color: #9b59b6;
                border-radius: 4px;
            }
        """)
        content_layout.addWidget(self.render_progress)
        
        content_layout.addWidget(QLabel(self.tr("当前步骤:")))
        self.step_label = QLabel(self.tr("等待开始"))
        self.step_label.setStyleSheet('color: #667eea; font-weight: bold;')
        content_layout.addWidget(self.step_label)
        
        content_layout.addWidget(QLabel(self.tr("总进度:")))
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        content_layout.addWidget(self.progress)
        
        self.process_info = QLabel()
        self.process_info.setStyleSheet('color: #888; font-size: 11px;')
        self.process_info.setAlignment(Qt.AlignCenter)
        content_layout.addWidget(self.process_info)
        
        self.start_btn = QPushButton(self.tr("▶ 开始测量"))
        self.start_btn.setStyleSheet('QPushButton { background-color: #667eea; color: white; font-weight: bold; font-size: 16px; padding: 12px; border-radius: 8px; } QPushButton:hover { background-color: #764ba2; }')
        self.start_btn.clicked.connect(self.start_measure)
        content_layout.addWidget(self.start_btn)

        # 停止测量 / 清空结果 按钮
        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        self.stop_btn = QPushButton(self.tr("⏹ 停止测量"))
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet('QPushButton { background-color: #c0392b; color: white; font-size: 12px; padding: 8px; border-radius: 6px; } QPushButton:hover { background-color: #e74c3c; } QPushButton:disabled { background-color: #555; color: #888; }')
        self.stop_btn.clicked.connect(self.stop_measure)
        btn_row.addWidget(self.stop_btn)

        self.clear_btn = QPushButton(self.tr("🧹 清空结果"))
        self.clear_btn.setStyleSheet('QPushButton { background-color: #7f8c8d; color: white; font-size: 12px; padding: 8px; border-radius: 6px; } QPushButton:hover { background-color: #95a5a6; }')
        self.clear_btn.clicked.connect(self.clear_results)
        btn_row.addWidget(self.clear_btn)
        content_layout.addLayout(btn_row)

        content_layout.addStretch()
        layout.addWidget(content_widget)
        layout.addStretch()
        return panel
    def _create_right_panel(self):
        panel = QGroupBox(self.tr("📊 结果与导出"))
        layout = QVBoxLayout(panel)
        layout.setSpacing(8)
        
        self.result_table = QTableWidget(5, 2)
        self.result_table.setHorizontalHeaderLabels([self.tr("指标"), self.tr("数值")])
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.result_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.result_table.setColumnWidth(0, 170)
        self.result_table.setColumnWidth(1, 90)
        
        metrics = [self.tr("节目响度(I)"), self.tr("最大短时响度(S)"), self.tr("最大瞬时响度(M)"), self.tr("最大真峰值(TP)"), self.tr("响度范围(LRA)")]
        for i, m in enumerate(metrics):
            self.result_table.setItem(i, 0, QTableWidgetItem(m))
            item = QTableWidgetItem("--")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.result_table.setItem(i, 1, item)
        
        layout.addWidget(self.result_table)
        
        status_layout = QHBoxLayout()
        self.int_status = QLabel(self.tr("节目响度: --"))
        self.int_status.setAlignment(Qt.AlignCenter)
        self.int_status.setStyleSheet("padding: 4px; border-radius: 4px;")
        status_layout.addWidget(self.int_status)
        
        self.tp_status = QLabel(self.tr("峰值: --"))
        self.tp_status.setAlignment(Qt.AlignCenter)
        self.tp_status.setStyleSheet("padding: 4px; border-radius: 4px;")
        status_layout.addWidget(self.tp_status)
        layout.addLayout(status_layout)
        
        export_box = QGroupBox(self.tr("导出"))
        export_layout = QVBoxLayout(export_box)
        
        btn_layout = QHBoxLayout()
        self.export_txt_btn = QPushButton("📄 TXT")
        self.export_txt_btn.setEnabled(False)
        self.export_txt_btn.clicked.connect(lambda: self.export_direct('txt'))
        btn_layout.addWidget(self.export_txt_btn)
        
        self.export_json_btn = QPushButton("📊 JSON")
        self.export_json_btn.setEnabled(False)
        self.export_json_btn.clicked.connect(lambda: self.export_direct('json'))
        btn_layout.addWidget(self.export_json_btn)
        
        self.export_excel_btn = QPushButton("📈 Excel")
        self.export_excel_btn.setEnabled(False)
        self.export_excel_btn.clicked.connect(lambda: self.export_direct('excel'))
        btn_layout.addWidget(self.export_excel_btn)
        export_layout.addLayout(btn_layout)
        
        export_note = QLabel(self.tr("TXT: 文本报告 | JSON: 结构化数据\nExcel: 包含每秒详细数据"))
        export_note.setStyleSheet("color: #888; font-size: 10px;")
        export_note.setAlignment(Qt.AlignCenter)
        export_layout.addWidget(export_note)
        
        layout.addWidget(export_box)

        # 短时响度/时间曲线图
        self.loudness_curve = LoudnessCurveWidget()
        layout.addWidget(self.loudness_curve)

        self.status = QLabel(self.tr("就绪"))
        self.status.setAlignment(Qt.AlignCenter)
        self.status.setStyleSheet("color: #888; font-size: 11px;")
        layout.addWidget(self.status)

        layout.addStretch()
        return panel
    
    def _apply_theme(self):
        self.setStyleSheet("""
            QMainWindow { background-color: rgba(26, 26, 46, 153); }
            QWidget { background-color: rgba(26, 26, 46, 153); color: #eee; }
            QGroupBox {
                background-color: rgba(26, 26, 46, 153);
                border: 2px solid #667eea;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 8px;
                font-weight: bold;
                font-size: 12px;
            }
            QGroupBox::title { color: #667eea; }
            QTableWidget {
                background-color: rgba(22, 33, 62, 153);
                border: 1px solid #667eea;
                font-size: 12px;
            }
            QHeaderView::section {
                background-color: #667eea;
                color: white;
                padding: 4px;
            }
            QPushButton {
                background-color: rgba(15, 52, 96, 153);
                border: 1px solid #667eea;
                padding: 6px 12px;
                border-radius: 4px;
            }
            QPushButton:hover { background-color: #667eea; }
            QComboBox {
                background-color: rgba(22, 33, 62, 153);
                border: 1px solid #667eea;
                padding: 4px;
            }
            QProgressBar {
                border: 1px solid #667eea;
                border-radius: 4px;
                text-align: center;
            }
            QProgressBar::chunk { background-color: #667eea; }
            QLabel { font-size: 12px; }
            QLineEdit {
                background-color: rgba(22, 33, 62, 153);
                border: 1px solid #667eea;
                padding: 4px;
            }
        """)
    
    def update_std_info(self):
        std = self.current_standard
        self.std_info.setText(
            self.tr("目标: {target} LKFS (±{tol} LU)\n峰值: {peak} dBTP")
            .format(target=f"{std.integrated_target:+.1f}",
                    tol=f"{std.integrated_tolerance:.1f}",
                    peak=f"{std.true_peak_limit:+.1f}")
        )
    
    def on_std_changed(self, text):
        # Reverse map: find original key from translated or raw text
        for key in LOUDNESS_STANDARDS:
            if self.tr(key) == text or key == text:
                self.current_standard = LOUDNESS_STANDARDS[key]
                self.update_std_info()
                break
    
    def on_input_mode_changed(self, mode):
        """输入模式切换"""
        self._update_mode_buttons(mode)
        self._update_mode_ui(mode)
        self.current_file = None
        self.mono_files = None
        self.current_adm_parser = None
        self.rendered_file = None

        if hasattr(self, 'filename_label'):
            self.filename_label.setText(self.tr("未选择文件"))
            self.filename_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #667eea;")
        if hasattr(self, 'path_label'):
            self.path_label.setText("")
        self._clear_file_metadata()
        if hasattr(self, 'channel_order_label'):
            self.channel_order_label.setText(self.tr("声道顺序: -"))

        if hasattr(self, 'adm_info'):
            self.adm_info.clear()
            self.adm_info.setPlaceholderText(self.tr("ADM文件信息将显示在这里..."))
        if hasattr(self, 'atmos_render_group'):
            self.atmos_render_group.setVisible(False)
        if hasattr(self, 'mono_files_group'):
            self.mono_files_group.setVisible(False)
        if hasattr(self, 'loudness_curve'):
            self.loudness_curve.set_data([], 0.0, -23.0)


    
    def _render_and_measure_adm(self):
        """渲染 ADM 文件到目标布局，然后测量响度（后台线程带进度条）"""
        if not self.current_file or not Path(self.current_file).exists():
            QMessageBox.warning(self, self.tr("提示"), self.tr("请先选择 ADM 文件"))
            return

        target_layout = self.atmos_layout_combo.currentText()

        from renderers.ear_renderer import is_object_based_adm

        if not is_object_based_adm(self.current_file):
            QMessageBox.information(self, self.tr("提示"), self.tr("该文件不包含动态对象音频，无需渲染。"))
            return

        # 禁用渲染按钮防止重复点击
        self.atmos_render_btn.setEnabled(False)

        # 显示渲染进度条
        self.render_step_label.setText(self.tr("🎧 正在渲染到 {layout}...").format(layout=target_layout))
        self.render_step_label.setVisible(True)
        self.render_progress.setValue(0)
        self.render_progress.setVisible(True)

        # 获取动态对象数量，用于初始化状态提示
        num_objects = 0
        if hasattr(self, 'current_adm_parser') and self.current_adm_parser and self.current_adm_parser.adm:
            adm = self.current_adm_parser.adm
            num_objects = len([ch for ch in adm.channel_formats if ch.type == 'Objects'])

        # 启动后台渲染线程
        # EAR 渲染初始化时会深度递归创建 point_source 对象，Mac 上 QThread
        # 默认栈大小（约 512KB）容易导致栈溢出，显式设置为 8MB。
        self._render_worker = ADMRenderWorker(self.current_file, target_layout, num_objects=num_objects)
        self._render_worker.setStackSize(8 * 1024 * 1024)
        self._render_worker.progress.connect(self._on_render_progress)
        self._render_worker.status.connect(self._on_render_status)
        self._render_worker.finished_signal.connect(self._on_render_finished)
        self._render_worker.error.connect(self._on_render_error)
        self._render_worker.start()

    def _on_render_progress(self, percent: int):
        """更新渲染进度条"""
        self.render_progress.setValue(percent)

    def _on_render_status(self, message: str):
        """更新渲染状态文本（如初始化提示）"""
        self.render_step_label.setText(message)

    def _on_render_finished(self, output_path: str):
        """渲染完成，保留 ADM UI 信息，仅更新文件信息为渲染后的文件"""
        self.render_step_label.setText(self.tr("🎧 渲染完成"))
        self.render_progress.setValue(100)
        self.atmos_render_btn.setEnabled(True)
        
        # 短暂等待，确保 EAR 写入的临时文件完全落盘后再开始测量
        # （Mac 上偶尔出现文件句柄/缓存未同步导致后续读取异常）
        import time
        time.sleep(0.5)

        # 保留 ADM 解析器与 UI，仅更新文件信息区域
        self.rendered_file = output_path
        self.current_file = output_path
        p = Path(output_path)
        self.filename_label.setText(self.tr("✓ 渲染: {name}").format(name=p.name))
        self.filename_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #9b59b6;")
        self.path_label.setText(str(p.parent))
        self._update_file_metadata(output_path)
        self._update_channel_order(output_path)

        # 自动设置声道配置
        cfg_map = {"Stereo (2.0)": "stereo", "5.1 (6ch)": "5.1", "7.1 (8ch)": "7.1",
                   "5.1.4 (10ch)": "5.1.4", "7.1.4 (12ch)": "7.1.4", "9.1.6 (16ch)": "9.1.6"}
        if self.atmos_layout_combo.currentText() in cfg_map:
            self.config_combo.setCurrentText(cfg_map[self.atmos_layout_combo.currentText()])

        # 自动开始测量
        self.start_measure()

    def _on_render_error(self, error_msg: str):
        """渲染出错"""
        self.render_step_label.setText(self.tr("🎧 渲染失败"))
        self.render_progress.setValue(0)
        self.render_progress.setVisible(False)
        self.atmos_render_btn.setEnabled(True)
        QMessageBox.critical(self, self.tr("渲染失败"), self.tr("渲染过程中出错:\n{err}").format(err=error_msg))

    def browse(self):
        """浏览文件（统一入口）"""
        files, _ = QFileDialog.getOpenFileNames(
            self, self.tr("选择音频文件"), "", "音频 (*.wav *.flac *.mp3 *.ogg)"
        )
        if not files:
            return

        if len(files) == 1:
            # 单个文件：判断是 ADM 还是普通多声道
            path = files[0]
            p = Path(path)
            try:
                info = sf.info(path)
            except Exception as e:
                QMessageBox.critical(self, self.tr("错误"), self.tr("无法读取文件: {err}").format(err=str(e)))
                return

            # 先重置状态，再检测 ADM（避免 on_input_mode_changed 清除 adm_parser）
            self.on_input_mode_changed('file')
            self.current_file = path
            self.mono_files = None
            self.filename_label.setText(f"✓ {p.name}")
            self.filename_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #27ae60;")
            self.path_label.setText(str(p.parent))
            self._update_file_metadata(path)
            self._update_channel_order(path)

            # 检查是否为 ADM/BW64（自动检测）
            if is_adm_file(path):
                self.current_adm_parser = BW64Parser(path)
                self.current_adm_parser.parse()
                self.parse_and_display_adm(path)
                self._update_mode_ui('file')  # 刷新 ADM 区域显隐

            cfg_map = {2: 'stereo', 6: '5.1', 8: '7.1', 10: '5.1.4', 12: '7.1.4'}
            if info.channels in cfg_map:
                self.config_combo.setCurrentText(cfg_map[info.channels])

        else:
            # 多个文件：强制要求所有文件都是单声道 WAV
            file_infos = {}
            for path in files:
                try:
                    info = sf.info(path)
                    file_infos[path] = info
                except Exception as e:
                    QMessageBox.critical(self, self.tr("错误"), self.tr("无法读取文件 {name}: {err}").format(name=Path(path).name, err=str(e)))
                    return
                if info.channels != 1:
                    QMessageBox.critical(
                        self, self.tr("声道错误"),
                        self.tr("文件 {name} 不是单声道（{ch} 声道）。\n\n多文件模式要求所有文件必须是单声道。").format(name=Path(path).name, ch=info.channels)
                    )
                    return
                if Path(path).suffix.lower() != '.wav':
                    QMessageBox.critical(
                        self, self.tr("格式错误"),
                        self.tr("文件 {name} 不是 WAV 格式。\n\n多文件模式仅支持 WAV。").format(name=Path(path).name)
                    )
                    return


            # 校验所有单声道文件时长是否一致
            durations = {p: file_infos[p].duration for p in files}
            unique_durations = sorted(set(durations.values()))
            if len(unique_durations) > 1:
                from collections import Counter
                duration_counts = Counter(durations.values())
                common_duration = duration_counts.most_common(1)[0][0]
                different = [(Path(p).name, d) for p, d in durations.items() if abs(d - common_duration) > 1e-6]
                
                lines = [self.tr("以下文件时长与其他文件不一致："), ""]
                for name, d in different:
                    lines.append(f"  • {name}: {self._format_duration(d)}")
                lines.append("")
                lines.append(self.tr("多数文件时长: {duration}").format(duration=self._format_duration(common_duration)))
                lines.append("")
                lines.append(self.tr("请统一所有文件时长后重新导入测量。"))
                
                QMessageBox.warning(
                    self, self.tr("时长不一致"),
                    "\\n".join(lines)
                )
                return

            common_duration = unique_durations[0]
            common_sr = file_infos[files[0]].samplerate
            common_subtype = file_infos[files[0]].subtype_info
            total_size = sum(Path(p).stat().st_size for p in files)
            # 进入多单声道模式
            self.on_input_mode_changed('file')
            self.current_file = None
            self.mono_files = [(p, '?') for p in files]
            self.filename_label.setText(self.tr("✓ {count} 个单声道文件").format(count=len(files)))
            self.filename_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #27ae60;")
            self.path_label.setText(str(Path(files[0]).parent))

            # 显示聚合后的文件元数据
            self.file_meta_labels['format'].setText('WAV')
            self.file_meta_labels['channels'].setText(str(len(files)))
            self.file_meta_labels['samplerate'].setText(f'{common_sr} Hz')
            self.file_meta_labels['bit_depth'].setText(str(common_subtype))
            self.file_meta_labels['duration'].setText(self._format_duration(common_duration))
            self.file_meta_labels['file_size'].setText(self._format_file_size(total_size))

            # 自动匹配声道
            template_map = {
                "Stereo (2.0)": "Stereo (2.0)",
                "5.1 (6ch)": "5.1 (6ch)",
                "7.1 (8ch)": "7.1 (8ch)",
                "7.1.2 (10ch)": "7.1.2 (10ch)",
                "5.1.4 (10ch)": "5.1.4 (10ch)",
                "7.1.4 (12ch)": "7.1.4 (12ch)",
            }
            combo_text = self.mono_template_combo.currentText()
            template_name = template_map.get(combo_text, None)
            self.mono_files = auto_match_mono_files(files, template_name=template_name)

            cfg_map = {2: 'stereo', 6: '5.1', 8: '7.1', 10: '5.1.4', 12: '7.1.4'}
            # 多单声道模式的通道数就是文件数量，必须按实际文件数设置配置，
            # 而不是按自动匹配成功的数量（可能有部分文件未能识别）。
            file_count = len(self.mono_files)
            if file_count in cfg_map:
                self.config_combo.setCurrentText(cfg_map[file_count])
            elif file_count > 0:
                # 文件数量不在标准配置中时回退到自动检测
                self.config_combo.setCurrentText(self.tr("自动检测"))

            self._update_mono_files_list()
            # 刷新 UI 显隐，确保 mono_section 正确显示
            self._update_mode_ui('file')

    def show_easy_import_help(self):
        """显示 EasyImport 功能说明"""
        QMessageBox.information(
            self,
            self.tr("EasyImport 说明"),
            self.tr(
                "EasyImport：\n"
                "直接选择任意封装格式待测文件，自动识别\n"
                "  ★ 多个单声道——自动完成文件->声道映射\n"
                "  ★ 单个多声道——自动识别内部顺序\n"
                "  ★ ADM BWF——解析元数据信息，提供目标声道格式选择及“渲染并测量”功能"
            ),
        )

    
    def start_measure(self):
        input_mode = None
        input_data = None

        # 优先判断多单声道模式
        if self.mono_files:
            input_mode = 'mono_list'
            input_data = self.mono_files
        elif hasattr(self, 'rendered_file') and self.rendered_file and self.current_file:
            # 渲染后的 ADM：测量渲染生成的文件，而非原始 ADM
            input_mode = 'file'
            input_data = self.rendered_file
        elif hasattr(self, 'current_adm_parser') and self.current_adm_parser and self.current_file:
            input_mode = 'adm'
            input_data = self.current_adm_parser
        elif self.current_file:
            input_mode = 'file'
            input_data = self.current_file
        else:
            QMessageBox.warning(self, self.tr("提示"), self.tr("请先选择输入文件"))
            return
        
        # 获取当前声道配置和标准
        config_name = self.config_combo.currentText()
        standard = self.current_standard
        
        # 创建并启动工作线程
        # 测量过程涉及大量 numpy/scipy 运算，增大线程栈大小避免 Mac 上栈溢出。
        self.worker = DetailedMeasurementWorker(input_mode, input_data, config_name, standard)
        self.worker.setStackSize(8 * 1024 * 1024)
        self.worker.sub_step.connect(self.on_sub_step)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.error.connect(self.on_error)
        
        # 重置UI状态
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress.setVisible(True)
        self.progress.setValue(0)
        self.export_txt_btn.setEnabled(False)
        self.export_json_btn.setEnabled(False)
        self.export_excel_btn.setEnabled(False)
        self.current_results = None
        
        # 启动线程
        self.worker.start()

    
    def on_sub_step(self, step, pct):
        self.step_label.setText(step)
        self.progress.setValue(pct)
        # 更新处理信息（包含倍速）
        if '⚡' in step:
            self.process_info.setText(step.split('|')[-1].strip())
    
    def on_progress(self, pct, current, total, speed, eta_str):
        self.progress.setValue(pct)
        if speed > 0:
            self.process_info.setText(self.tr("{current:.1f}s/{total:.1f}s | ⚡{speed:.1f}x 实时 | {eta}")
                .format(current=current, total=total, speed=speed, eta=eta_str))
        else:
            self.process_info.setText(f"{current:.1f}s/{total:.1f}s | {eta_str}")
    
    def on_finished(self, results):
        # 合并文件信息到结果
        results['file_info'] = self._build_file_info()
        self.current_results = results
        self.progress.setVisible(False)
        self.start_btn.setEnabled(True)
        self.step_label.setText(self.tr("完成"))
        
        std = self.current_standard
        
        # 更新结果表格
        data = [
            (f"{results['integrated']:+.2f}", "LKFS"),
            (f"{results['short_term']:+.2f}", "LKFS"),
            (f"{results['momentary']:+.2f}", "LKFS"),
            (f"{results['true_peak']:+.2f}", "dBTP"),
            (f"{results['lra']:.2f}", "LU")
        ]
        
        for i, (val, unit) in enumerate(data):
            item = QTableWidgetItem(f"{val} {unit}")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.result_table.setItem(i, 1, item)
        
        # 更新合规状态
        int_ok = std.check_integrated(results['integrated'])
        tp_ok = std.check_true_peak(results['true_peak'])
        
        self.int_status.setText(self.tr("节目响度: {status}").format(status=('✓' if int_ok else '✗')))
        self.int_status.setStyleSheet(f"background-color: {'#27ae60' if int_ok else '#e74c3c'}; color: white; padding: 4px; border-radius: 4px;")
        
        self.tp_status.setText(self.tr("峰值: {status}").format(status=('✓' if tp_ok else '✗')))
        self.tp_status.setStyleSheet(f"background-color: {'#27ae60' if tp_ok else '#e74c3c'}; color: white; padding: 4px; border-radius: 4px;")
        
        self.status.setText(self.tr("完成 | 用时 {time:.1f}s").format(time=results.get('processing_time', 0)))
        self.export_txt_btn.setEnabled(True)
        self.export_json_btn.setEnabled(True)
        self.export_excel_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

        # 更新短时响度曲线图
        curve_data = results.get('short_term_curve', [])
        duration = results.get('duration', 0.0)
        if curve_data and duration > 0:
            self.loudness_curve.set_data(curve_data, duration, std.integrated_target)

    def stop_measure(self):
        """停止当前测量线程"""
        if self.worker and self.worker.isRunning():
            self.worker.requestInterruption()
            if not self.worker.wait(3000):
                self.worker.terminate()
            self.step_label.setText(self.tr("已停止"))
            self.status.setText(self.tr("测量已停止"))
            self.progress.setVisible(False)
            self.start_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)

    def clear_results(self):
        """清空右侧结果数据"""
        self.current_results = None
        for i in range(self.result_table.rowCount()):
            item = QTableWidgetItem("--")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.result_table.setItem(i, 1, item)
        self.int_status.setText(self.tr("节目响度: --"))
        self.int_status.setStyleSheet("padding: 4px; border-radius: 4px;")
        self.tp_status.setText(self.tr("峰值: --"))
        self.tp_status.setStyleSheet("padding: 4px; border-radius: 4px;")
        self.status.setText(self.tr("就绪"))
        self.export_txt_btn.setEnabled(False)
        self.export_json_btn.setEnabled(False)
        self.export_excel_btn.setEnabled(False)
        if hasattr(self, 'loudness_curve'):
            self.loudness_curve.set_data([], 0.0, -23.0)

    def on_error(self, msg):
        self.progress.setVisible(False)
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.step_label.setText(self.tr("错误"))
        QMessageBox.critical(self, self.tr("错误"), msg)

    def resizeEvent(self, event):
        """窗口大小变化时更新背景层尺寸"""
        super().resizeEvent(event)
        if hasattr(self, 'bg_label') and self.bg_label and self.centralWidget():
            self.bg_label.setGeometry(self.centralWidget().rect())
    
    def _build_file_info(self) -> dict:
        """构建当前被测文件信息字典"""
        info = {
            'mode': 'unknown',
            'file_path': '',
            'file_name': '',
            'adm_info': '',
            'renderer_info': '',
            'authoring_info': '',
            'ref_layout': '',
            'mono_files': []
        }
        
        if self.mono_files:
            info['mode'] = 'mono'
            info['file_path'] = str(Path(self.mono_files[0][0]).parent)
            info['file_name'] = self.tr("{count} 个单声道文件").format(count=len(self.mono_files))
            info['mono_files'] = [
                {'path': path, 'channel': ch, 'name': Path(path).name}
                for path, ch in self.mono_files
            ]
        elif hasattr(self, 'current_adm_parser') and self.current_adm_parser and self.current_file:
            info['mode'] = 'adm'
            info['file_path'] = str(Path(self.current_file).parent)
            info['file_name'] = Path(self.current_file).name
            if hasattr(self, 'adm_info'):
                info['adm_info'] = self.adm_info.toPlainText()
        elif self.current_file:
            info['mode'] = 'file'
            info['file_path'] = str(Path(self.current_file).parent)
            info['file_name'] = Path(self.current_file).name
        
        return info
    
    def _get_export_base_name(self) -> str:
        """生成导出文件名基础部分（取自被测文件共有名）"""
        if self.mono_files:
            stems = [Path(p).stem for p, _ in self.mono_files]
        elif self.current_file:
            return Path(self.current_file).stem
            if not stems:
                return "report"
            # 求共有前缀
            common = stems[0]
            for s in stems[1:]:
                i = 0
                while i < len(common) and i < len(s) and common[i] == s[i]:
                    i += 1
                common = common[:i]
            common = common.rstrip('._-')
            if common:
                return common
            return stems[0]
        return "report"

    def export_direct(self, fmt):
        if not self.current_results:
            return
        
        # 清理标准名中的 Windows 非法字符
        std_name = self.current_standard.name.replace(' ', '_').replace('/', '_').replace('\\', '_')[:15]
        ext_map = {'txt': 'txt', 'json': 'json', 'excel': 'xlsx'}
        ext = ext_map.get(fmt, fmt)
        base = self._get_export_base_name()
        now = datetime.now()
        timestamp = f"{now:%y%m%d}_{now:%H}H{now:%M}"
        default_name = f"{base}_{timestamp}_{std_name}.{ext}"
        
        filter_map = {
            'txt': self.tr("文本文件 (*.txt)"),
            'json': self.tr("JSON文件 (*.json)"),
            'excel': self.tr("Excel文件 (*.xlsx)")
        }
        filter_str = filter_map.get(fmt, f"*.{ext}")
        
        # 使用 QFileDialog 类确保默认文件名在 Windows 原生对话框中正确显示
        dialog = QFileDialog(self, self.tr("导出"))
        dialog.setAcceptMode(QFileDialog.AcceptSave)
        dialog.setNameFilter(filter_str)
        dialog.selectFile(default_name)
        if self.current_file:
            dialog.setDirectory(str(Path(self.current_file).parent))
        if dialog.exec() != QFileDialog.Accepted:
            return
        path = dialog.selectedFiles()[0]
        if not path:
            return
        
        try:
            fi = self.current_results.get('file_info', {})
            if fmt == 'excel':
                self._export_excel_detailed(path, {})
            else:
                results_obj = LoudnessResults(
                    integrated=self.current_results['integrated'],
                    short_term=self.current_results['short_term'],
                    momentary=self.current_results['momentary'],
                    true_peak=self.current_results['true_peak'],
                    lra=self.current_results['lra'],
                    max_true_peak=self.current_results['max_true_peak'],
                    duration=self.current_results['duration'],
                    filename=self.current_results.get('filename', 'unknown'),
                    sample_rate=self.current_results['sample_rate'],
                    channels=self.current_results['channels'],
                    blocks=[],
                    file_path=fi.get('file_path', ''),
                    file_name=fi.get('file_name', ''),
                    adm_info=fi.get('adm_info', ''),
                    renderer_info=fi.get('renderer_info', ''),
                    authoring_info=fi.get('authoring_info', ''),
                    ref_layout=fi.get('ref_layout', ''),
                    mono_files=fi.get('mono_files', [])
                )
                exporter = ReportExporter(results_obj)
                
                if fmt == 'txt':
                    exporter.export_txt(path)
                elif fmt == 'json':
                    exporter.export_json(path)
            
            self.status.setText(self.tr("已导出: {name}").format(name=Path(path).name))
            
        except Exception as e:
            QMessageBox.critical(self, self.tr("导出失败"), self.tr("导出失败: {err}").format(err=str(e)))
    
    def _export_excel_detailed(self, path: str, opts: dict):
        """导出详细Excel报告 (.xlsx)
        
        使用 openpyxl 生成真正的 Excel 文件，实现：
        - 中文正确显示（微软雅黑字体）
        - 超标数值用红色粗体 + 浅红色背景标注
        - 专业的表格格式
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(self, self.tr("缺少依赖"), self.tr("请安装 openpyxl: pip install openpyxl"))
            return
        
        detailed = self.current_results.get('detailed_data')
        standard = self.current_standard
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.tr("响度测量报告")
        
        # 样式定义
        title_font = Font(name='微软雅黑', size=14, bold=True, color='2F5496')
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        normal_font = Font(name='微软雅黑', size=10)
        red_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
        exceed_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        row = 1
        
        # === 被测文件信息 ===
        fi = self.current_results.get('file_info', {})
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1, value=self.tr('被测文件信息'))
        cell.font = title_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        row += 2
        
        file_info_rows = [
            [self.tr('测量时间'), time.strftime('%Y-%m-%d %H:%M:%S')],
            [self.tr('文件路径'), fi.get('file_path', '-')],
            [self.tr('文件名称'), fi.get('file_name', self.current_results.get('filename', 'unknown'))],
        ]
        if fi.get('renderer_info'):
            file_info_rows.append([self.tr('渲染器'), fi['renderer_info']])
        if fi.get('authoring_info'):
            file_info_rows.append([self.tr('创作软件'), fi['authoring_info']])
        if fi.get('ref_layout'):
            file_info_rows.append([self.tr('参考布局'), fi['ref_layout']])
        if fi.get('mono_files'):
            for item in fi['mono_files']:
                file_info_rows.append([self.tr("声道 {channel}").format(channel=item.get('channel', '?')), item.get('name', '')])
        
        for label, value in file_info_rows:
            ws.cell(row=row, column=1, value=label).font = normal_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row, column=2, value=value).font = normal_font
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        
        # ADM 信息单独处理：保留换行，合并单元格并自动换行显示
        if fi.get('adm_info'):
            ws.cell(row=row, column=1, value=self.tr('ADM 信息')).font = normal_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='top')
            adm_cell = ws.cell(row=row, column=2, value=fi['adm_info'])
            adm_cell.font = Font(name='Consolas', size=9)
            adm_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            line_count = fi['adm_info'].count('\n') + 1
            ws.row_dimensions[row].height = max(30, line_count * 14)
            row += 1
        
        row += 1
        
        # === 整体测量结果 ===
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1, value=self.tr('整体测量结果'))
        cell.font = title_font
        cell.alignment = center_align
        row += 2
        
        headers = [self.tr('指标'), self.tr('数值'), self.tr('单位'), self.tr('标准限值'), self.tr('状态')]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1
        
        summary = detailed['summary'] if detailed else {
            'integrated_loudness': self.current_results['integrated'],
            'max_short_term': self.current_results['short_term'],
            'max_momentary': self.current_results['momentary'],
            'max_true_peak': self.current_results['true_peak'],
            'lra': self.current_results['lra'],
            'duration': self.current_results['duration']
        }
        
        # 节目响度
        int_val = summary['integrated_loudness']
        int_target = standard.integrated_target
        int_tol = standard.integrated_tolerance
        int_status = self.tr('合格') if abs(int_val - int_target) <= int_tol else self.tr('超标')
        is_exceed = int_status == '超标'
        ws.cell(row=row, column=1, value=self.tr('节目响度')).font = normal_font
        ws.cell(row=row, column=2, value=f'{int_val:+.2f}').font = red_font if is_exceed else normal_font
        ws.cell(row=row, column=3, value='LKFS').font = normal_font
        ws.cell(row=row, column=4, value=f'{int_target:+.1f} ± {int_tol:.1f}').font = normal_font
        status_cell = ws.cell(row=row, column=5, value=int_status)
        status_cell.font = red_font if is_exceed else normal_font
        if is_exceed:
            status_cell.fill = exceed_fill
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align
        row += 1
        
        # 最大短时响度
        st_val = summary['max_short_term']
        ws.cell(row=row, column=1, value=self.tr('最大短时响度')).font = normal_font
        ws.cell(row=row, column=2, value=f'{st_val:+.2f}').font = normal_font
        ws.cell(row=row, column=3, value='LKFS').font = normal_font
        ws.cell(row=row, column=4, value='-').font = normal_font
        ws.cell(row=row, column=5, value=self.tr('合格')).font = normal_font
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align
        row += 1
        
        # 最大瞬时响度
        mom_val = summary['max_momentary']
        ws.cell(row=row, column=1, value=self.tr('最大瞬时响度')).font = normal_font
        ws.cell(row=row, column=2, value=f'{mom_val:+.2f}').font = normal_font
        ws.cell(row=row, column=3, value='LKFS').font = normal_font
        ws.cell(row=row, column=4, value='-').font = normal_font
        ws.cell(row=row, column=5, value=self.tr('合格')).font = normal_font
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align
        row += 1
        
        # 最大真峰值
        tp_val = summary['max_true_peak']
        tp_limit = standard.true_peak_limit
        tp_status = self.tr('合格') if tp_val <= tp_limit else self.tr('超标')
        is_exceed = tp_status == '超标'
        ws.cell(row=row, column=1, value=self.tr('最大真峰值')).font = normal_font
        ws.cell(row=row, column=2, value=f'{tp_val:+.2f}').font = red_font if is_exceed else normal_font
        ws.cell(row=row, column=3, value='dBTP').font = normal_font
        ws.cell(row=row, column=4, value=f'≤ {tp_limit:+.1f}').font = normal_font
        status_cell = ws.cell(row=row, column=5, value=tp_status)
        status_cell.font = red_font if is_exceed else normal_font
        if is_exceed:
            status_cell.fill = exceed_fill
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align
        row += 1
        
        # 响度范围
        lra_val = summary['lra']
        ws.cell(row=row, column=1, value=self.tr('响度范围(LRA)')).font = normal_font
        ws.cell(row=row, column=2, value=f'{lra_val:.2f}').font = normal_font
        ws.cell(row=row, column=3, value='LU').font = normal_font
        ws.cell(row=row, column=4, value='-').font = normal_font
        ws.cell(row=row, column=5, value=self.tr('参考')).font = normal_font
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align
        row += 1
        
        # 时长
        ws.cell(row=row, column=1, value=self.tr('测量时长')).font = normal_font
        ws.cell(row=row, column=2, value=f"{summary['duration']:.2f}").font = normal_font
        ws.cell(row=row, column=3, value=self.tr('秒')).font = normal_font
        ws.cell(row=row, column=4, value='-').font = normal_font
        ws.cell(row=row, column=5, value='-').font = normal_font
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align
        row += 1
        
        row += 2
        
        # === 每秒短时响度 ===
        if detailed and detailed.get('short_term_per_second'):
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws.cell(row=row, column=1, value=self.tr('每秒短时响度 (3秒滑动窗口)'))
            cell.font = title_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 2
            
            st_headers = [self.tr('时间(秒)'), self.tr('短时响度(LKFS)'), self.tr('状态')]
            for col, header in enumerate(st_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            row += 1
            
            for item in detailed['short_term_per_second']:
                time_sec = item['time']
                lufs = item['lufs']
                is_exceed = item.get('is_exceed', False)
                
                ws.cell(row=row, column=1, value=f'{time_sec:.1f}').font = normal_font
                ws.cell(row=row, column=1).alignment = center_align
                ws.cell(row=row, column=1).border = thin_border
                
                lufs_cell = ws.cell(row=row, column=2, value=f'{lufs:+.2f}')
                lufs_cell.font = red_font if is_exceed else normal_font
                lufs_cell.alignment = center_align
                lufs_cell.border = thin_border
                if is_exceed:
                    lufs_cell.fill = exceed_fill
                
                status_cell = ws.cell(row=row, column=3, value=self.tr('超标') if is_exceed else self.tr('合格'))
                status_cell.font = red_font if is_exceed else normal_font
                status_cell.alignment = center_align
                status_cell.border = thin_border
                if is_exceed:
                    status_cell.fill = exceed_fill
                
                row += 1
            
            row += 1
        
        # === 每秒最大真峰值 ===
        if detailed and detailed.get('true_peak_per_second'):
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws.cell(row=row, column=1, value=self.tr('每秒最大真峰值'))
            cell.font = title_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 2
            
            tp_headers = [self.tr('时间(秒)'), self.tr('真峰值(dBTP)'), self.tr('标准限值'), self.tr('状态')]
            for col, header in enumerate(tp_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            row += 1
            
            for item in detailed['true_peak_per_second']:
                time_sec = item['time']
                dbtp = item['dbtp']
                is_exceed = item.get('is_exceed', False)
                
                ws.cell(row=row, column=1, value=f'{time_sec:.1f}').font = normal_font
                ws.cell(row=row, column=1).alignment = center_align
                ws.cell(row=row, column=1).border = thin_border
                
                dbtp_cell = ws.cell(row=row, column=2, value=f'{dbtp:+.2f}')
                dbtp_cell.font = red_font if is_exceed else normal_font
                dbtp_cell.alignment = center_align
                dbtp_cell.border = thin_border
                if is_exceed:
                    dbtp_cell.fill = exceed_fill
                
                ws.cell(row=row, column=3, value=f'≤ {tp_limit:+.1f}').font = normal_font
                ws.cell(row=row, column=3).alignment = center_align
                ws.cell(row=row, column=3).border = thin_border
                
                status_cell = ws.cell(row=row, column=4, value=self.tr('超标') if is_exceed else self.tr('合格'))
                status_cell.font = red_font if is_exceed else normal_font
                status_cell.alignment = center_align
                status_cell.border = thin_border
                if is_exceed:
                    status_cell.fill = exceed_fill
                
                row += 1
            
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        
        wb.save(path)
        print(f'[Excel导出] 已保存到: {path}')


def _get_resource_path(relative_path):
    """获取 PyInstaller 打包后的资源路径（兼容 onefile / onedir / 开发环境）"""
    if hasattr(sys, '_MEIPASS'):
        # onefile 模式：资源在临时解压目录
        return Path(sys._MEIPASS) / relative_path
    # onedir 模式或开发环境：优先尝试 EXE 所在目录
    exe_dir = Path(sys.executable).parent
    candidate = exe_dir / relative_path
    if candidate.exists():
        return candidate
    # 开发环境回退
    return Path(__file__).parent.parent / relative_path


def _show_splash(app):
    """显示启动画面，覆盖程序初始化时间"""
    from PySide6.QtWidgets import QSplashScreen
    from PySide6.QtGui import QPixmap, QPainter, QFont, QColor, QFontMetrics
    from PySide6.QtCore import Qt

    # 加载背景图
    bg_path = _get_resource_path("assets/bg.png")
    bg_pixmap = QPixmap(str(bg_path))
    if bg_pixmap.isNull():
        # 回退：纯色背景
        bg_pixmap = QPixmap(800, 600)
        bg_pixmap.fill(QColor("#0f0f23"))

    # Splash 窗口尺寸（只比卡片大一小圈）
    splash_w, splash_h = 440, 300
    card_w, card_h = 420, 280

    # 截取背景图左下角 440×300
    src_x = 0
    src_y = max(0, bg_pixmap.height() - splash_h)
    bg_source = bg_pixmap.copy(src_x, src_y, min(splash_w, bg_pixmap.width()), min(splash_h, bg_pixmap.height()))

    pixmap = QPixmap(splash_w, splash_h)
    painter = QPainter(pixmap)
    painter.drawPixmap(0, 0, bg_source)

    # 卡片居中
    card_x = (splash_w - card_w) // 2
    card_y = (splash_h - card_h) // 2

    # 60% 透明卡片背景
    painter.setPen(Qt.NoPen)
    painter.setBrush(QColor(26, 26, 46, 153))
    painter.drawRoundedRect(card_x, card_y, card_w, card_h, 12, 12)

    # 卡片边框
    painter.setPen(QColor("#667eea"))
    painter.setBrush(Qt.NoBrush)
    painter.drawRoundedRect(card_x, card_y, card_w, card_h, 12, 12)

    # Logo 图片
    logo_path = str(_get_resource_path("assets/centerlogo.png"))
    logo = QPixmap(logo_path)
    if not logo.isNull():
        # 缩放至适合卡片宽度（最大 300px 宽）
        logo_w = min(300, logo.width())
        logo_h = int(logo.height() * (logo_w / logo.width()))
        logo = logo.scaled(logo_w, logo_h, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        logo_x = card_x + (card_w - logo_w) // 2
        logo_y = card_y + 20
        painter.drawPixmap(logo_x, logo_y, logo)
    else:
        # 回退：文字
        painter.setPen(QColor("#667eea"))
        painter.setFont(QFont("Segoe UI", 56, QFont.Bold))
        painter.drawText(card_x + (card_w - 60) // 2, card_y + 100, "LD")

    # 版本号
    painter.setPen(QColor("#667eea"))
    painter.setFont(QFont("Segoe UI", 10))
    fm = QFontMetrics(painter.font())
    text = "v1.0.3  (build 260630)"
    x = card_x + (card_w - fm.horizontalAdvance(text)) // 2
    painter.drawText(x, card_y + 175, text)

    # © 2026 YOYH All Rights Reserved
    painter.setPen(QColor("#8899cc"))
    painter.setFont(QFont("Segoe UI", 9))
    fm = QFontMetrics(painter.font())
    text = "\u00a9 2026 YOYH All Rights Reserved"
    x = card_x + (card_w - fm.horizontalAdvance(text)) // 2
    painter.drawText(x, card_y + 205, text)

    # Loading...
    painter.setPen(QColor("#8899cc"))
    painter.setFont(QFont("Segoe UI", 9))
    fm = QFontMetrics(painter.font())
    text = "Loading..."
    x = card_x + (card_w - fm.horizontalAdvance(text)) // 2
    painter.drawText(x, card_y + 250, text)

    painter.end()

    splash = QSplashScreen(pixmap, Qt.WindowStaysOnTopHint | Qt.FramelessWindowHint)
    splash.show()
    app.processEvents()
    return splash


def ensure_single_instance(app):
    """单实例守护：已有实例在运行时激活它并退出当前进程"""
    from PySide6.QtNetwork import QLocalServer, QLocalSocket
    from PySide6.QtWidgets import QApplication, QMainWindow

    socket_name = "LouDuck_SingleInstance"

    # 尝试连接已有实例
    socket = QLocalSocket()
    socket.connectToServer(socket_name)
    if socket.waitForConnected(500):
        socket.write(b"activate")
        socket.flush()
        socket.waitForBytesWritten(500)
        socket.disconnectFromServer()
        return False

    # 无已有实例，创建本地服务器
    server = QLocalServer()
    server.removeServer(socket_name)
    if not server.listen(socket_name):
        return False

    # 防止被垃圾回收
    app._single_instance_server = server

    def on_new_connection():
        conn = server.nextPendingConnection()
        if conn and conn.waitForReadyRead(500):
            data = conn.readAll().data()
            if data == b"activate":
                for widget in QApplication.topLevelWidgets():
                    if isinstance(widget, QMainWindow):
                        widget.showNormal()
                        widget.raise_()
                        widget.activateWindow()
                        break

    server.newConnection.connect(on_new_connection)
    return True


def main():
    import argparse
    import faulthandler
    # 启用 faulthandler，在 SIGBUS/SIGSEGV 等 C 层面崩溃时输出 traceback，
    # 便于定位 Mac 上打包版闪退的根因。
    faulthandler.enable()
    
    parser = argparse.ArgumentParser()
    parser.add_argument('--lang', default=None, help='Force language, e.g. en, zh')
    args, remaining = parser.parse_known_args()
    
    # 在导入 numpy/scipy 之前设置 BLAS 单线程，避免 Apple Silicon 上
    # 多线程并发导致的 SIGBUS / bus error（常见于 EAR 渲染时）。
    if sys.platform == 'darwin':
        os.environ.setdefault('OPENBLAS_NUM_THREADS', '1')
        os.environ.setdefault('VECLIB_MAXIMUM_THREADS', '1')
        os.environ.setdefault('MKL_NUM_THREADS', '1')
        os.environ.setdefault('NUMEXPR_NUM_THREADS', '1')
    
    app = QApplication([sys.argv[0]] + remaining)
    app.setStyle('Fusion')

    # 单实例检查
    if not ensure_single_instance(app):
        print("[INFO] Another instance is already running. Activating it.")
        sys.exit(0)

    # 显示启动画面（必须在重模块导入之前，让用户立刻看到反馈）
    splash = _show_splash(app)
    
    # 延迟导入重模块，减少启动前等待时间
    global sf, np, ITU1770Meter, ChannelConfig, LoudnessResults, ReportExporter, BW64Parser, is_adm_file
    import soundfile as sf
    import numpy as np
    from itu1770_meter import ITU1770Meter, ChannelConfig
    from report_exporter import LoudnessResults, ReportExporter
    from adm_parser import BW64Parser, is_adm_file
    
    # i18n: auto-detect system locale and load translation if available
    from PySide6.QtCore import QTranslator, QLocale
    translator = QTranslator()
    
    if args.lang:
        # 强制指定语言
        qm_path = _get_resource_path(f'i18n/LouDuck_{args.lang}.qm')
        if not qm_path.exists():
            # 尝试加上地区后缀
            qm_path = _get_resource_path(f'i18n/LouDuck_{args.lang}_US.qm')
    else:
        # 跨平台读取系统显示语言
        locale_name = QLocale.system().name()
        try:
            if sys.platform == 'win32':
                import ctypes
                import locale as py_locale
                lang_id = ctypes.windll.kernel32.GetUserDefaultUILanguage()
                locale_name = py_locale.windows_locale.get(lang_id, locale_name)
            elif sys.platform == 'darwin':
                # macOS: 读取 AppleLanguages
                import subprocess
                import ast
                result = subprocess.run(
                    ['defaults', 'read', '-g', 'AppleLanguages'],
                    capture_output=True, text=True, check=True
                )
                langs = ast.literal_eval(result.stdout.strip())
                if langs:
                    locale_name = langs[0].replace('-', '_')
        except Exception:
            pass
        qm_path = _get_resource_path(f'i18n/LouDuck_{locale_name}.qm')
        if not qm_path.exists() and '_' in locale_name:
            # fallback to language code only, e.g. "en" from "en_US"
            lang = locale_name.split('_')[0]
            qm_path = _get_resource_path(f'i18n/LouDuck_{lang}.qm')
    
    if qm_path.exists():
        if translator.load(str(qm_path)):
            app.installTranslator(translator)
    
    icon_path = _get_resource_path('assets/icon.ico')
    if icon_path.exists():
        from PySide6.QtGui import QIcon
        app.setWindowIcon(QIcon(str(icon_path)))
    win = LoudnessMeterApp()
    splash.finish(win)
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
